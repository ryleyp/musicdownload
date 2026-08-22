from __future__ import annotations

import base64
import csv
import hashlib
import json
import os
import secrets
import shutil
import sqlite3
import tempfile
import time
import urllib.parse
import webbrowser
from contextlib import contextmanager
from datetime import datetime, timezone
from http.server import BaseHTTPRequestHandler, HTTPServer
from pathlib import Path
from typing import Any, Iterable

try:
    import requests
except ImportError:  # Report a project-specific setup message at runtime.
    requests = None  # type: ignore[assignment]


MODULE_DIR = Path(__file__).resolve().parent
PROJECT_OVERRIDE = os.environ.get("MUSIC_LIBRARY_PROJECT_DIR", "").strip()
CURRENT_DIR = Path.cwd().resolve()
if PROJECT_OVERRIDE:
    PROJECT_DIR = Path(PROJECT_OVERRIDE).expanduser().resolve()
elif (CURRENT_DIR / ".env").exists() or (CURRENT_DIR / "data").is_dir():
    PROJECT_DIR = CURRENT_DIR
else:
    PROJECT_DIR = MODULE_DIR
DEFAULT_DATA_DIR = PROJECT_DIR / "data"
DEFAULT_DB_PATH = DEFAULT_DATA_DIR / "music_library.sqlite"
DEFAULT_CSV_PATH = DEFAULT_DATA_DIR / "music_library.csv"
DEFAULT_XLSX_PATH = DEFAULT_DATA_DIR / "music_library_review.xlsx"
DEFAULT_REDIRECT_URI = "http://127.0.0.1:8888/callback"
SPOTIFY_API = "https://api.spotify.com/v1"
SPOTIFY_ACCOUNTS = "https://accounts.spotify.com"


class AppError(RuntimeError):
    pass


def running_project_venv() -> bool:
    expected = PROJECT_DIR / ".venv"
    try:
        return Path(os.sys.prefix).resolve() == expected.resolve()
    except (OSError, ValueError):
        return False


def dependency_help(package: str) -> str:
    activation = ""
    if not running_project_venv():
        activation = (
            f" The project virtual environment is not active. Run: "
            f"cd {PROJECT_DIR!s} && source .venv/bin/activate."
        )
    return (
        f"{package} is not available to {os.sys.executable}.{activation} "
        "Then run: python -m pip install -r requirements.txt"
    )


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def load_dotenv(path: Path | None = None) -> None:
    env_path = path or PROJECT_DIR / ".env"
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip("\"'")
        if key and key not in os.environ:
            os.environ[key] = value


def ensure_data_dir(db_path: Path) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)


@contextmanager
def connect_db(db_path: Path = DEFAULT_DB_PATH):
    ensure_data_dir(db_path)
    connection = sqlite3.connect(db_path)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA busy_timeout = 60000")
    connection.execute("PRAGMA foreign_keys = ON")
    connection.execute("PRAGMA journal_mode = WAL")
    initialize_schema(connection)
    try:
        yield connection
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def initialize_schema(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        CREATE TABLE IF NOT EXISTS tracks (
            spotify_id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            artists TEXT NOT NULL,
            artist_ids TEXT,
            primary_artist TEXT,
            primary_artist_id TEXT,
            album TEXT,
            album_id TEXT,
            album_artist TEXT,
            release_date TEXT,
            release_year INTEGER,
            disc_number INTEGER,
            track_number INTEGER,
            total_tracks INTEGER,
            duration_ms INTEGER,
            explicit INTEGER,
            isrc TEXT,
            spotify_url TEXT,
            album_url TEXT,
            cover_url TEXT,
            genres TEXT,
            added_at TEXT,
            is_liked INTEGER NOT NULL DEFAULT 1,
            is_saved_album INTEGER NOT NULL DEFAULT 0,
            user_deleted INTEGER NOT NULL DEFAULT 0,
            youtube_url TEXT,
            youtube_video_id TEXT,
            youtube_title TEXT,
            youtube_channel TEXT,
            youtube_channel_verified INTEGER NOT NULL DEFAULT 0,
            youtube_duration_seconds INTEGER,
            youtube_score REAL,
            youtube_score_notes TEXT,
            youtube_score_version INTEGER NOT NULL DEFAULT 1,
            match_status TEXT NOT NULL DEFAULT 'not_searched',
            match_attempts INTEGER NOT NULL DEFAULT 0,
            last_match_attempt_at TEXT,
            match_error TEXT,
            match_error_code TEXT,
            match_next_retry_at TEXT,
            reviewed_at TEXT,
            download_status TEXT NOT NULL DEFAULT 'not_downloaded',
            download_path TEXT,
            download_error TEXT,
            download_error_code TEXT,
            download_next_retry_at TEXT,
            download_attempts INTEGER NOT NULL DEFAULT 0,
            last_download_attempt_at TEXT,
            downloaded_at TEXT,
            downloaded_duration_seconds REAL,
            downloaded_duration_difference_seconds REAL,
            apple_music_status TEXT NOT NULL DEFAULT 'not_checked',
            apple_music_preferred_id TEXT,
            apple_music_replaced_ids TEXT,
            apple_music_updated_at TEXT,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS artist_cache (
            spotify_artist_id TEXT PRIMARY KEY,
            artist_name TEXT NOT NULL,
            genres TEXT,
            fetch_status TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS youtube_candidates (
            spotify_id TEXT NOT NULL,
            youtube_video_id TEXT NOT NULL,
            rank INTEGER NOT NULL,
            youtube_url TEXT NOT NULL,
            youtube_title TEXT,
            youtube_channel TEXT,
            youtube_channel_verified INTEGER NOT NULL DEFAULT 0,
            youtube_duration_seconds INTEGER,
            source_tier INTEGER NOT NULL DEFAULT 0,
            score REAL NOT NULL,
            score_notes TEXT,
            hard_reject INTEGER NOT NULL DEFAULT 0,
            PRIMARY KEY (spotify_id, youtube_video_id),
            FOREIGN KEY (spotify_id) REFERENCES tracks(spotify_id)
                ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS match_assessments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id TEXT NOT NULL,
            spotify_id TEXT NOT NULL,
            youtube_video_id TEXT,
            youtube_url TEXT,
            youtube_title TEXT,
            youtube_channel TEXT,
            youtube_duration_seconds REAL,
            score REAL,
            hard_reject INTEGER NOT NULL DEFAULT 0,
            automatic_approval_eligible INTEGER NOT NULL DEFAULT 0,
            approval_threshold REAL,
            decision TEXT NOT NULL,
            reason TEXT,
            score_notes TEXT,
            scoring_version INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY (spotify_id) REFERENCES tracks(spotify_id)
        );

        CREATE TABLE IF NOT EXISTS workflow_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            spotify_id TEXT,
            stage TEXT NOT NULL,
            event TEXT NOT NULL,
            status TEXT NOT NULL,
            details_json TEXT NOT NULL DEFAULT '{}',
            created_at TEXT NOT NULL,
            FOREIGN KEY (spotify_id) REFERENCES tracks(spotify_id)
        );

        CREATE TABLE IF NOT EXISTS workflow_state (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS apple_music_restore_plans (
            plan_id TEXT PRIMARY KEY,
            spotify_id TEXT NOT NULL,
            action TEXT NOT NULL,
            download_path TEXT NOT NULL,
            playlist_name TEXT NOT NULL,
            old_tracks_json TEXT NOT NULL,
            preferred_id TEXT,
            preferred_location TEXT,
            preferred_existed_before INTEGER,
            preferred_enabled_before INTEGER,
            playlist_had_track INTEGER,
            status TEXT NOT NULL,
            created_at TEXT NOT NULL,
            applied_at TEXT,
            restored_at TEXT,
            restore_error TEXT,
            FOREIGN KEY (spotify_id) REFERENCES tracks(spotify_id)
        );

        CREATE TABLE IF NOT EXISTS spotify_playlist_sync_runs (
            run_id TEXT PRIMARY KEY,
            synced_at TEXT NOT NULL,
            playlist_count INTEGER NOT NULL,
            item_count INTEGER NOT NULL
        );

        CREATE TABLE IF NOT EXISTS spotify_playlists (
            run_id TEXT NOT NULL,
            spotify_playlist_id TEXT NOT NULL,
            name TEXT NOT NULL,
            description TEXT,
            owner_name TEXT,
            spotify_url TEXT,
            snapshot_id TEXT,
            item_count INTEGER NOT NULL,
            fetch_error TEXT,
            PRIMARY KEY (run_id, spotify_playlist_id),
            FOREIGN KEY (run_id) REFERENCES spotify_playlist_sync_runs(run_id)
        );

        CREATE TABLE IF NOT EXISTS spotify_playlist_tracks (
            run_id TEXT NOT NULL,
            spotify_playlist_id TEXT NOT NULL,
            position INTEGER NOT NULL,
            spotify_track_id TEXT,
            title TEXT,
            artist TEXT,
            album TEXT,
            added_at TEXT,
            item_type TEXT,
            is_local INTEGER NOT NULL DEFAULT 0,
            PRIMARY KEY (run_id, spotify_playlist_id, position),
            FOREIGN KEY (run_id, spotify_playlist_id)
                REFERENCES spotify_playlists(run_id, spotify_playlist_id)
        );

        CREATE TABLE IF NOT EXISTS music_genre_cache (
            cache_key TEXT PRIMARY KEY,
            album_artist TEXT NOT NULL,
            album TEXT NOT NULL,
            genre TEXT,
            source TEXT NOT NULL,
            status TEXT NOT NULL,
            catalog_url TEXT,
            error TEXT,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS music_genre_runs (
            run_id TEXT PRIMARY KEY,
            mode TEXT NOT NULL,
            status TEXT NOT NULL,
            planned_count INTEGER NOT NULL,
            applied_count INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            completed_at TEXT
        );

        CREATE TABLE IF NOT EXISTS music_genre_changes (
            run_id TEXT NOT NULL,
            music_persistent_id TEXT NOT NULL,
            old_genre TEXT NOT NULL,
            new_genre TEXT NOT NULL,
            source TEXT NOT NULL,
            status TEXT NOT NULL,
            error TEXT,
            PRIMARY KEY (run_id, music_persistent_id),
            FOREIGN KEY (run_id) REFERENCES music_genre_runs(run_id)
        );

        CREATE TABLE IF NOT EXISTS music_metadata_runs (
            run_id TEXT PRIMARY KEY,
            status TEXT NOT NULL,
            planned_count INTEGER NOT NULL,
            applied_count INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            completed_at TEXT
        );

        CREATE TABLE IF NOT EXISTS music_metadata_changes (
            run_id TEXT NOT NULL,
            music_persistent_id TEXT NOT NULL,
            spotify_id TEXT NOT NULL,
            old_values_json TEXT NOT NULL,
            new_values_json TEXT NOT NULL,
            status TEXT NOT NULL,
            error TEXT,
            PRIMARY KEY (run_id, music_persistent_id),
            FOREIGN KEY (run_id) REFERENCES music_metadata_runs(run_id),
            FOREIGN KEY (spotify_id) REFERENCES tracks(spotify_id)
        );

        CREATE TABLE IF NOT EXISTS music_delete_runs (
            run_id TEXT PRIMARY KEY,
            playlist_name TEXT NOT NULL,
            status TEXT NOT NULL,
            planned_count INTEGER NOT NULL,
            deleted_count INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            completed_at TEXT
        );

        CREATE TABLE IF NOT EXISTS music_delete_items (
            run_id TEXT NOT NULL,
            music_persistent_id TEXT NOT NULL,
            spotify_id TEXT,
            title TEXT,
            artist TEXT,
            album TEXT,
            music_location TEXT,
            download_path TEXT,
            status TEXT NOT NULL,
            error TEXT,
            PRIMARY KEY (run_id, music_persistent_id),
            FOREIGN KEY (run_id) REFERENCES music_delete_runs(run_id)
        );

        CREATE TABLE IF NOT EXISTS music_album_artist_runs (
            run_id TEXT PRIMARY KEY,
            scope TEXT NOT NULL,
            status TEXT NOT NULL,
            planned_count INTEGER NOT NULL,
            applied_count INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            completed_at TEXT
        );

        CREATE TABLE IF NOT EXISTS music_album_artist_changes (
            run_id TEXT NOT NULL,
            music_persistent_id TEXT NOT NULL,
            title TEXT,
            album TEXT,
            old_album_artist TEXT NOT NULL,
            new_album_artist TEXT NOT NULL,
            old_compilation INTEGER NOT NULL,
            new_compilation INTEGER NOT NULL,
            status TEXT NOT NULL,
            error TEXT,
            PRIMARY KEY (run_id, music_persistent_id),
            FOREIGN KEY (run_id) REFERENCES music_album_artist_runs(run_id)
        );

        CREATE TABLE IF NOT EXISTS music_group_cleanup_runs (
            run_id TEXT PRIMARY KEY,
            status TEXT NOT NULL,
            group_count INTEGER NOT NULL,
            planned_count INTEGER NOT NULL,
            applied_count INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            completed_at TEXT
        );

        CREATE TABLE IF NOT EXISTS music_group_cleanup_changes (
            run_id TEXT NOT NULL,
            music_persistent_id TEXT NOT NULL,
            title TEXT,
            track_artist TEXT,
            old_album TEXT NOT NULL,
            new_album TEXT NOT NULL,
            old_album_artist TEXT NOT NULL,
            new_album_artist TEXT NOT NULL,
            old_compilation INTEGER NOT NULL,
            new_compilation INTEGER NOT NULL,
            status TEXT NOT NULL,
            error TEXT,
            PRIMARY KEY (run_id, music_persistent_id),
            FOREIGN KEY (run_id) REFERENCES music_group_cleanup_runs(run_id)
        );

        CREATE TABLE IF NOT EXISTS music_artist_credit_runs (
            run_id TEXT PRIMARY KEY,
            status TEXT NOT NULL,
            group_count INTEGER NOT NULL,
            planned_count INTEGER NOT NULL,
            applied_count INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            completed_at TEXT
        );

        CREATE TABLE IF NOT EXISTS music_artist_credit_changes (
            run_id TEXT NOT NULL,
            music_persistent_id TEXT NOT NULL,
            title TEXT,
            album TEXT,
            old_artist TEXT NOT NULL,
            new_artist TEXT NOT NULL,
            old_album_artist TEXT NOT NULL,
            new_album_artist TEXT NOT NULL,
            old_sort_artist TEXT NOT NULL,
            new_sort_artist TEXT NOT NULL,
            old_sort_album_artist TEXT NOT NULL,
            new_sort_album_artist TEXT NOT NULL,
            status TEXT NOT NULL,
            error TEXT,
            PRIMARY KEY (run_id, music_persistent_id),
            FOREIGN KEY (run_id) REFERENCES music_artist_credit_runs(run_id)
        );

        CREATE INDEX IF NOT EXISTS idx_tracks_match_status
            ON tracks(match_status);
        CREATE INDEX IF NOT EXISTS idx_tracks_is_liked
            ON tracks(is_liked);
        CREATE INDEX IF NOT EXISTS idx_candidates_spotify_rank
            ON youtube_candidates(spotify_id, rank);
        CREATE INDEX IF NOT EXISTS idx_assessments_spotify_created
            ON match_assessments(spotify_id, created_at);
        CREATE INDEX IF NOT EXISTS idx_events_stage_created
            ON workflow_events(stage, created_at);
        CREATE INDEX IF NOT EXISTS idx_restore_plans_status
            ON apple_music_restore_plans(status, created_at);
        CREATE INDEX IF NOT EXISTS idx_spotify_playlist_runs_synced
            ON spotify_playlist_sync_runs(synced_at);
        CREATE INDEX IF NOT EXISTS idx_spotify_playlist_tracks_spotify
            ON spotify_playlist_tracks(spotify_track_id);
        CREATE INDEX IF NOT EXISTS idx_music_genre_changes_status
            ON music_genre_changes(run_id, status);
        CREATE INDEX IF NOT EXISTS idx_music_metadata_changes_status
            ON music_metadata_changes(run_id, status);
        CREATE INDEX IF NOT EXISTS idx_music_delete_items_status
            ON music_delete_items(run_id, status);
        CREATE INDEX IF NOT EXISTS idx_music_album_artist_changes_status
            ON music_album_artist_changes(run_id, status);
        CREATE INDEX IF NOT EXISTS idx_music_group_cleanup_changes_status
            ON music_group_cleanup_changes(run_id, status);
        """
    )
    existing_columns = {
        row[1] for row in connection.execute("PRAGMA table_info(tracks)").fetchall()
    }
    migrations = {
        "download_attempts": (
            "ALTER TABLE tracks ADD COLUMN "
            "download_attempts INTEGER NOT NULL DEFAULT 0"
        ),
        "last_download_attempt_at": (
            "ALTER TABLE tracks ADD COLUMN last_download_attempt_at TEXT"
        ),
        "downloaded_at": "ALTER TABLE tracks ADD COLUMN downloaded_at TEXT",
        "apple_music_status": (
            "ALTER TABLE tracks ADD COLUMN "
            "apple_music_status TEXT NOT NULL DEFAULT 'not_checked'"
        ),
        "apple_music_preferred_id": (
            "ALTER TABLE tracks ADD COLUMN apple_music_preferred_id TEXT"
        ),
        "apple_music_replaced_ids": (
            "ALTER TABLE tracks ADD COLUMN apple_music_replaced_ids TEXT"
        ),
        "apple_music_updated_at": (
            "ALTER TABLE tracks ADD COLUMN apple_music_updated_at TEXT"
        ),
        "youtube_channel_verified": (
            "ALTER TABLE tracks ADD COLUMN "
            "youtube_channel_verified INTEGER NOT NULL DEFAULT 0"
        ),
        "is_saved_album": (
            "ALTER TABLE tracks ADD COLUMN "
            "is_saved_album INTEGER NOT NULL DEFAULT 0"
        ),
        "user_deleted": (
            "ALTER TABLE tracks ADD COLUMN "
            "user_deleted INTEGER NOT NULL DEFAULT 0"
        ),
        "youtube_score_version": (
            "ALTER TABLE tracks ADD COLUMN "
            "youtube_score_version INTEGER NOT NULL DEFAULT 1"
        ),
        "match_attempts": (
            "ALTER TABLE tracks ADD COLUMN match_attempts INTEGER NOT NULL DEFAULT 0"
        ),
        "last_match_attempt_at": (
            "ALTER TABLE tracks ADD COLUMN last_match_attempt_at TEXT"
        ),
        "match_error": "ALTER TABLE tracks ADD COLUMN match_error TEXT",
        "match_error_code": (
            "ALTER TABLE tracks ADD COLUMN match_error_code TEXT"
        ),
        "match_next_retry_at": (
            "ALTER TABLE tracks ADD COLUMN match_next_retry_at TEXT"
        ),
        "download_error_code": (
            "ALTER TABLE tracks ADD COLUMN download_error_code TEXT"
        ),
        "download_next_retry_at": (
            "ALTER TABLE tracks ADD COLUMN download_next_retry_at TEXT"
        ),
        "downloaded_duration_seconds": (
            "ALTER TABLE tracks ADD COLUMN downloaded_duration_seconds REAL"
        ),
        "downloaded_duration_difference_seconds": (
            "ALTER TABLE tracks ADD COLUMN "
            "downloaded_duration_difference_seconds REAL"
        ),
    }
    for column, statement in migrations.items():
        if column not in existing_columns:
            connection.execute(statement)
    connection.execute(
        "CREATE INDEX IF NOT EXISTS idx_tracks_is_saved_album "
        "ON tracks(is_saved_album)"
    )
    connection.execute(
        "CREATE INDEX IF NOT EXISTS idx_tracks_user_deleted "
        "ON tracks(user_deleted)"
    )
    candidate_columns = {
        row[1]
        for row in connection.execute(
            "PRAGMA table_info(youtube_candidates)"
        ).fetchall()
    }
    if "youtube_channel_verified" not in candidate_columns:
        connection.execute(
            "ALTER TABLE youtube_candidates ADD COLUMN "
            "youtube_channel_verified INTEGER NOT NULL DEFAULT 0"
        )
    if "source_tier" not in candidate_columns:
        connection.execute(
            "ALTER TABLE youtube_candidates ADD COLUMN "
            "source_tier INTEGER NOT NULL DEFAULT 0"
        )


def parse_release_year(release_date: str | None) -> int | None:
    if not release_date:
        return None
    try:
        return int(release_date[:4])
    except (TypeError, ValueError):
        return None


def duration_text(duration_ms: int | None) -> str:
    if duration_ms is None:
        return ""
    seconds = round(duration_ms / 1000)
    return f"{seconds // 60}:{seconds % 60:02d}"


def duration_seconds_text(seconds: int | float | None) -> str:
    if seconds is None:
        return ""
    rounded = round(float(seconds))
    return f"{rounded // 60}:{rounded % 60:02d}"


class SpotifyClient:
    def __init__(
        self,
        data_dir: Path,
        *,
        scopes: tuple[str, ...] = ("user-library-read",),
        token_filename: str = ".spotify_token.json",
    ):
        if requests is None:
            raise AppError(dependency_help("requests"))
        load_dotenv()
        self.client_id = os.environ.get("SPOTIFY_CLIENT_ID", "").strip()
        self.redirect_uri = os.environ.get(
            "SPOTIFY_REDIRECT_URI", DEFAULT_REDIRECT_URI
        ).strip()
        if not self.client_id or self.client_id == "paste_your_client_id_here":
            raise AppError(
                "SPOTIFY_CLIENT_ID is missing. Copy .env.example to .env and "
                "paste the Client ID from your Spotify developer app."
            )
        self.scopes = tuple(dict.fromkeys(scopes))
        self.token_path = data_dir / token_filename
        self.session = requests.Session()
        self.token = self._load_token()

    def _load_token(self) -> dict[str, Any] | None:
        if not self.token_path.exists():
            return None
        try:
            return json.loads(self.token_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return None

    def _save_token(self, token: dict[str, Any]) -> None:
        self.token_path.parent.mkdir(parents=True, exist_ok=True)
        self.token_path.write_text(
            json.dumps(token, indent=2, sort_keys=True), encoding="utf-8"
        )
        try:
            self.token_path.chmod(0o600)
        except OSError:
            pass
        self.token = token

    def _token_expired(self) -> bool:
        if not self.token:
            return True
        return float(self.token.get("expires_at", 0)) <= time.time() + 60

    def access_token(self) -> str:
        if not self.token:
            self._authorize_pkce()
        elif self._token_expired():
            try:
                self._refresh_token()
            except AppError:
                self.token = None
                self._authorize_pkce()
        assert self.token is not None
        return str(self.token["access_token"])

    def _authorize_pkce(self) -> None:
        parsed = urllib.parse.urlparse(self.redirect_uri)
        if parsed.scheme != "http" or parsed.hostname != "127.0.0.1":
            raise AppError(
                "For this local Mac tool, SPOTIFY_REDIRECT_URI must use an "
                "explicit loopback address such as "
                "http://127.0.0.1:8888/callback."
            )
        if not parsed.port:
            raise AppError("SPOTIFY_REDIRECT_URI must include a port.")

        verifier = secrets.token_urlsafe(72)[:96]
        challenge = base64.urlsafe_b64encode(
            hashlib.sha256(verifier.encode("ascii")).digest()
        ).decode("ascii").rstrip("=")
        state = secrets.token_urlsafe(24)
        scope = " ".join(self.scopes)
        auth_url = f"{SPOTIFY_ACCOUNTS}/authorize?" + urllib.parse.urlencode(
            {
                "client_id": self.client_id,
                "response_type": "code",
                "redirect_uri": self.redirect_uri,
                "scope": scope,
                "state": state,
                "code_challenge_method": "S256",
                "code_challenge": challenge,
            }
        )

        result: dict[str, str] = {}
        expected_path = parsed.path or "/"

        class CallbackHandler(BaseHTTPRequestHandler):
            def do_GET(self) -> None:
                request = urllib.parse.urlparse(self.path)
                params = urllib.parse.parse_qs(request.query)
                if request.path != expected_path:
                    self.send_response(404)
                    self.end_headers()
                    return
                result["state"] = params.get("state", [""])[0]
                result["code"] = params.get("code", [""])[0]
                result["error"] = params.get("error", [""])[0]
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(
                    b"<h2>Spotify connected.</h2>"
                    b"<p>You can close this tab and return to Terminal.</p>"
                )

            def log_message(self, *_args: Any) -> None:
                return

        try:
            server = HTTPServer(("127.0.0.1", parsed.port), CallbackHandler)
        except OSError as exc:
            raise AppError(
                f"Could not listen on 127.0.0.1:{parsed.port}. "
                "Close any app using that port and try again."
            ) from exc
        print("Opening Spotify in your browser for permission...")
        webbrowser.open(auth_url)
        server.timeout = 300
        server.handle_request()
        server.server_close()

        if result.get("error"):
            raise AppError(f"Spotify authorization was denied: {result['error']}")
        if not result.get("code") or result.get("state") != state:
            raise AppError(
                "Spotify authorization did not complete or the security state "
                "did not match. Run the command again."
            )

        try:
            response = self.session.post(
                f"{SPOTIFY_ACCOUNTS}/api/token",
                data={
                    "client_id": self.client_id,
                    "grant_type": "authorization_code",
                    "code": result["code"],
                    "redirect_uri": self.redirect_uri,
                    "code_verifier": verifier,
                },
                timeout=30,
            )
        except requests.RequestException as exc:
            raise AppError(
                "Could not reach Spotify during authorization. Check the "
                f"internet connection and try again: {exc}"
            ) from exc
        if not response.ok:
            raise AppError(
                f"Spotify token request failed ({response.status_code}): "
                f"{response.text[:300]}"
            )
        token = response.json()
        token["expires_at"] = time.time() + int(token.get("expires_in", 3600))
        self._save_token(token)

    def _refresh_token(self) -> None:
        assert self.token is not None
        refresh_token = self.token.get("refresh_token")
        if not refresh_token:
            raise AppError("No Spotify refresh token is available.")
        response = self.session.post(
            f"{SPOTIFY_ACCOUNTS}/api/token",
            data={
                "grant_type": "refresh_token",
                "refresh_token": refresh_token,
                "client_id": self.client_id,
            },
            timeout=30,
        )
        if not response.ok:
            raise AppError(
                f"Spotify token refresh failed ({response.status_code})."
            )
        refreshed = response.json()
        refreshed["refresh_token"] = refreshed.get(
            "refresh_token", refresh_token
        )
        refreshed["expires_at"] = time.time() + int(
            refreshed.get("expires_in", 3600)
        )
        self._save_token(refreshed)

    def get(
        self, endpoint_or_url: str, params: dict[str, Any] | None = None
    ) -> dict[str, Any]:
        url = (
            endpoint_or_url
            if endpoint_or_url.startswith("http")
            else f"{SPOTIFY_API}{endpoint_or_url}"
        )
        refreshed_once = False
        for attempt in range(7):
            try:
                response = self.session.get(
                    url,
                    params=params,
                    headers={"Authorization": f"Bearer {self.access_token()}"},
                    timeout=30,
                )
            except requests.RequestException as exc:
                if attempt < 2:
                    time.sleep(2 ** attempt)
                    continue
                raise AppError(
                    "Could not reach Spotify after several attempts. Your "
                    f"checkpoint is safe: {exc}"
                ) from exc
            if response.status_code == 401 and not refreshed_once:
                refreshed_once = True
                self._refresh_token()
                continue
            if response.status_code == 429:
                retry_after = response.headers.get("Retry-After")
                if retry_after is None:
                    reason = ""
                    try:
                        reason = response.json().get("error", {}).get("reason", "")
                    except ValueError:
                        pass
                    if reason == "QUOTA_EXCEEDED":
                        raise AppError(
                            "Spotify development quota was reached. Your saved "
                            "progress is safe. Try the command again later."
                        )
                    wait_seconds = min(60, 2 ** attempt)
                else:
                    wait_seconds = max(1, int(float(retry_after)))
                print(f"Spotify asked us to pause for {wait_seconds} seconds...")
                time.sleep(wait_seconds)
                continue
            if response.status_code >= 500:
                time.sleep(min(30, 2 ** attempt))
                continue
            if not response.ok:
                if response.status_code == 403:
                    raise AppError(
                        "Spotify returned 403 Forbidden. Confirm that the app "
                        "owner/account is eligible, the account is allowed in "
                        "the app dashboard, and these scopes were granted: "
                        + ", ".join(self.scopes)
                    )
                raise AppError(
                    f"Spotify API request failed ({response.status_code}): "
                    f"{response.text[:300]}"
                )
            return response.json()
        raise AppError("Spotify API did not recover after several retries.")


TRACK_EXPORT_COLUMNS = [
    "decision",
    "manual_youtube_url",
    "match_status",
    "library_source",
    "local_deleted_blocked",
    "confidence",
    "title",
    "artists",
    "album",
    "release_date",
    "year",
    "genres",
    "duration",
    "spotify_runtime",
    "spotify_duration_seconds",
    "explicit",
    "track_number",
    "disc_number",
    "isrc",
    "spotify_url",
    "youtube_url",
    "youtube_title",
    "youtube_channel",
    "youtube_runtime",
    "youtube_duration_seconds",
    "runtime_difference_seconds",
    "youtube_score",
    "youtube_score_version",
    "match_error",
    "match_error_code",
    "match_next_retry_at",
    "download_status",
    "download_error_code",
    "download_next_retry_at",
    "download_attempts",
    "downloaded_runtime",
    "downloaded_duration_seconds",
    "downloaded_duration_difference_seconds",
    "downloaded_at",
    "apple_music_status",
    "added_at",
    "spotify_id",
]


def track_export_rows(connection: sqlite3.Connection) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT *
        FROM tracks
        WHERE is_liked = 1 OR is_saved_album = 1
        ORDER BY primary_artist COLLATE NOCASE,
                 album COLLATE NOCASE,
                 disc_number,
                 track_number,
                 title COLLATE NOCASE
        """
    ).fetchall()
    output: list[dict[str, Any]] = []
    for row in rows:
        score = row["youtube_score"]
        confidence = ""
        if score is not None:
            confidence = (
                "high" if score >= 95 else "medium" if score >= 75 else "low"
            )
        output.append(
            {
                "decision": "",
                "manual_youtube_url": "",
                "match_status": row["match_status"],
                "library_source": "; ".join(
                    source
                    for source, included in (
                        ("liked_song", row["is_liked"]),
                        ("saved_album", row["is_saved_album"]),
                    )
                    if included
                ),
                "local_deleted_blocked": "Yes" if row["user_deleted"] else "No",
                "confidence": confidence,
                "title": row["title"],
                "artists": row["artists"],
                "album": row["album"],
                "release_date": row["release_date"],
                "year": row["release_year"],
                "genres": row["genres"],
                "duration": duration_text(row["duration_ms"]),
                "spotify_runtime": duration_text(row["duration_ms"]),
                "spotify_duration_seconds": (
                    round(row["duration_ms"] / 1000)
                    if row["duration_ms"] is not None
                    else None
                ),
                "explicit": "Yes" if row["explicit"] else "No",
                "track_number": row["track_number"],
                "disc_number": row["disc_number"],
                "isrc": row["isrc"],
                "spotify_url": row["spotify_url"],
                "youtube_url": row["youtube_url"],
                "youtube_title": row["youtube_title"],
                "youtube_channel": row["youtube_channel"],
                "youtube_runtime": duration_seconds_text(
                    row["youtube_duration_seconds"]
                ),
                "youtube_duration_seconds": row["youtube_duration_seconds"],
                "runtime_difference_seconds": (
                    abs(
                        round(row["duration_ms"] / 1000)
                        - row["youtube_duration_seconds"]
                    )
                    if row["duration_ms"] is not None
                    and row["youtube_duration_seconds"] is not None
                    else None
                ),
                "youtube_score": row["youtube_score"],
                "youtube_score_version": row["youtube_score_version"],
                "match_error": row["match_error"],
                "match_error_code": row["match_error_code"],
                "match_next_retry_at": row["match_next_retry_at"],
                "download_status": row["download_status"],
                "download_error_code": row["download_error_code"],
                "download_next_retry_at": row["download_next_retry_at"],
                "download_attempts": row["download_attempts"],
                "downloaded_runtime": duration_seconds_text(
                    row["downloaded_duration_seconds"]
                ),
                "downloaded_duration_seconds": row[
                    "downloaded_duration_seconds"
                ],
                "downloaded_duration_difference_seconds": row[
                    "downloaded_duration_difference_seconds"
                ],
                "downloaded_at": row["downloaded_at"],
                "apple_music_status": row["apple_music_status"],
                "added_at": row["added_at"],
                "spotify_id": row["spotify_id"],
            }
        )
    return output


def export_catalog(
    connection: sqlite3.Connection,
    csv_path: Path = DEFAULT_CSV_PATH,
    xlsx_path: Path = DEFAULT_XLSX_PATH,
) -> None:
    rows = track_export_rows(connection)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    backup_existing_file(csv_path)
    backup_existing_file(xlsx_path)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{csv_path.name}.", suffix=".tmp", dir=csv_path.parent
    )
    os.close(descriptor)
    temporary_csv = Path(temporary_name)
    try:
        with temporary_csv.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.DictWriter(handle, fieldnames=TRACK_EXPORT_COLUMNS)
            writer.writeheader()
            writer.writerows(rows)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary_csv, csv_path)
    finally:
        temporary_csv.unlink(missing_ok=True)
    export_workbook(connection, rows, xlsx_path)


def backup_existing_file(path: Path) -> Path | None:
    if not path.exists():
        return None
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ")
    backup_dir = path.parent / "backups" / timestamp
    backup_dir.mkdir(parents=True, exist_ok=False)
    backup_path = backup_dir / path.name
    shutil.copy2(path, backup_path)
    return backup_path


def export_workbook(
    connection: sqlite3.Connection,
    track_rows: list[dict[str, Any]],
    path: Path,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise AppError(
            "openpyxl is not installed. Run: python -m pip install -r "
            "requirements.txt"
        ) from exc

    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instruction_lines = [
        ["Spotify to YouTube Music Library"],
        [
            "1. Run spotify_sync.py to refresh liked songs, or add "
            "--include-albums to include tracks from saved albums."
        ],
        ["2. Run youtube_match.py to create ranked YouTube suggestions."],
        [
            "3. In Tracks, set decision to approve or skip. You may paste a "
            "better source in manual_youtube_url."
        ],
        ["4. Save and close this workbook, then run import_review.py."],
        ["5. Run download_mp3.py. Only approved rows are downloaded."],
        [
            "Preferred sources are artist Topic uploads, official audio, and "
            "official lyric videos. Music videos and altered versions are "
            "rejected unless Spotify names that exact version. Automatic "
            "approval requires a trusted source and runtime within 5 seconds."
        ],
        [
            "Do not run spotify_sync.py or youtube_match.py while you have "
            "unsaved decisions in this workbook."
        ],
    ]
    for line in instruction_lines:
        instructions.append(line)
    instructions["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    instructions["A1"].fill = PatternFill("solid", fgColor="1DB954")
    instructions.column_dimensions["A"].width = 110
    for row in instructions.iter_rows():
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    tracks_sheet = workbook.create_sheet("Tracks")
    tracks_sheet.append(TRACK_EXPORT_COLUMNS)
    for item in track_rows:
        tracks_sheet.append([item.get(column, "") for column in TRACK_EXPORT_COLUMNS])
    style_sheet(tracks_sheet, TRACK_EXPORT_COLUMNS)
    decision_validation = DataValidation(
        type="list", formula1='"approve,skip,reset"', allow_blank=True
    )
    tracks_sheet.add_data_validation(decision_validation)
    if tracks_sheet.max_row >= 2:
        decision_validation.add(f"A2:A{tracks_sheet.max_row}")
    add_hyperlinks(tracks_sheet, TRACK_EXPORT_COLUMNS)

    candidate_columns = [
        "spotify_id",
        "spotify_title",
        "artists",
        "spotify_runtime",
        "spotify_duration_seconds",
        "rank",
        "source_tier",
        "score",
        "hard_reject",
        "youtube_runtime",
        "youtube_duration_seconds",
        "duration_difference_seconds",
        "youtube_url",
        "youtube_title",
        "youtube_channel",
        "score_notes",
    ]
    candidates_sheet = workbook.create_sheet("Candidates")
    candidates_sheet.append(candidate_columns)
    candidate_rows = connection.execute(
        """
        SELECT c.*, t.title AS spotify_title, t.artists,
               t.duration_ms,
               CASE
                 WHEN c.youtube_duration_seconds IS NULL
                      OR t.duration_ms IS NULL THEN NULL
                 ELSE ABS(c.youtube_duration_seconds - ROUND(t.duration_ms / 1000.0))
               END AS duration_difference_seconds
        FROM youtube_candidates c
        JOIN tracks t USING (spotify_id)
        WHERE t.is_liked = 1 OR t.is_saved_album = 1
        ORDER BY t.primary_artist COLLATE NOCASE,
                 t.title COLLATE NOCASE,
                 c.rank
        """
    ).fetchall()
    for row in candidate_rows:
        item = dict(row)
        item["spotify_runtime"] = duration_text(row["duration_ms"])
        item["spotify_duration_seconds"] = (
            round(row["duration_ms"] / 1000)
            if row["duration_ms"] is not None
            else None
        )
        item["youtube_runtime"] = duration_seconds_text(
            row["youtube_duration_seconds"]
        )
        candidates_sheet.append([item[column] for column in candidate_columns])
    style_sheet(candidates_sheet, candidate_columns)
    add_hyperlinks(candidates_sheet, candidate_columns)

    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".xlsx", dir=path.parent
    )
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, path)
    finally:
        temporary_path.unlink(missing_ok=True)


def style_sheet(sheet: Any, columns: Iterable[str]) -> None:
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill("solid", fgColor="191414")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "decision": 13,
        "manual_youtube_url": 38,
        "match_status": 18,
        "library_source": 24,
        "confidence": 12,
        "title": 34,
        "spotify_title": 34,
        "artists": 28,
        "album": 32,
        "genres": 30,
        "spotify_url": 38,
        "youtube_url": 38,
        "youtube_title": 48,
        "youtube_channel": 28,
        "download_status": 18,
        "downloaded_at": 22,
        "apple_music_status": 23,
        "score_notes": 70,
        "match_error": 60,
        "match_error_code": 24,
        "download_error_code": 24,
        "added_at": 22,
        "spotify_id": 24,
    }
    for index, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(
            column, 16
        )
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)


def add_hyperlinks(sheet: Any, columns: list[str]) -> None:
    from openpyxl.styles import Font

    for name in ("spotify_url", "youtube_url", "manual_youtube_url"):
        if name not in columns:
            continue
        column_index = columns.index(name) + 1
        for row_index in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single")


def chunks(items: list[str], size: int) -> Iterable[list[str]]:
    for index in range(0, len(items), size):
        yield items[index : index + size]
