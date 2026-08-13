#!/usr/bin/env python3
"""Review, score, and download a focused cohort of recent Spotify additions."""
from __future__ import annotations

import argparse
import csv
import os
import subprocess
import sys
import tempfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from common import (
    AppError,
    DEFAULT_DB_PATH,
    PROJECT_DIR,
    backup_existing_file,
    connect_db,
    dependency_help,
)


DEFAULT_REPORT = PROJECT_DIR / "data" / "recent_spotify_additions.csv"
DEFAULT_XLSX = PROJECT_DIR / "data" / "recent_spotify_additions.xlsx"
REPORT_FIELDS = [
    "spotify_id", "added_at", "library_source", "title", "artists", "album",
    "spotify_runtime_seconds", "match_status", "youtube_score", "youtube_title",
    "youtube_channel", "youtube_runtime_seconds", "runtime_difference_seconds",
    "download_status", "review_state", "spotify_url", "youtube_url",
]


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Create a focused review of recently liked songs and saved-album "
            "tracks, optionally sync, score, and download only that cohort."
        )
    )
    parser.add_argument("--db", type=Path, default=DEFAULT_DB_PATH)
    parser.add_argument("--days", type=int, default=30)
    parser.add_argument(
        "--since",
        help="Use additions on/after YYYY-MM-DD instead of --days.",
    )
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument(
        "--sync",
        action="store_true",
        help="Refresh liked songs and saved albums from Spotify first.",
    )
    parser.add_argument(
        "--match",
        action="store_true",
        help="Score recent tracks that still need YouTube matching.",
    )
    parser.add_argument(
        "--refresh",
        action="store_true",
        help="With --match, re-search every recent track, including existing matches.",
    )
    parser.add_argument("--auto-approve", type=float, default=95.0)
    parser.add_argument(
        "--download",
        action="store_true",
        help="Download only eligible tracks from this recent cohort.",
    )
    parser.add_argument("--min-score", type=float, default=95.0)
    parser.add_argument("--retry-errors", action="store_true")
    parser.add_argument("--download-dry-run", action="store_true")
    parser.add_argument("--po-token-provider", action="store_true")
    parser.add_argument("--open-review", action="store_true")
    return parser.parse_args(argv)


def cutoff_text(args: argparse.Namespace) -> str:
    if args.days < 0:
        raise AppError("--days cannot be negative.")
    if args.since:
        try:
            parsed = datetime.strptime(args.since, "%Y-%m-%d").replace(
                tzinfo=timezone.utc
            )
        except ValueError as exc:
            raise AppError("--since must use YYYY-MM-DD.") from exc
        return parsed.isoformat()
    return (datetime.now(timezone.utc) - timedelta(days=args.days)).isoformat()


def recent_rows(connection: Any, cutoff: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for track in connection.execute(
        """
        SELECT * FROM tracks
        WHERE (is_liked = 1 OR is_saved_album = 1)
          AND user_deleted = 0
          AND added_at IS NOT NULL
          AND added_at >= ?
        ORDER BY added_at DESC, primary_artist COLLATE NOCASE, title COLLATE NOCASE
        """,
        (cutoff,),
    ):
        spotify_seconds = (
            round(float(track["duration_ms"]) / 1000.0, 3)
            if track["duration_ms"] is not None else ""
        )
        youtube_seconds = track["youtube_duration_seconds"]
        difference = (
            round(abs(float(youtube_seconds) - float(spotify_seconds)), 3)
            if youtube_seconds is not None and spotify_seconds != "" else ""
        )
        if track["download_status"] == "downloaded":
            review_state = "downloaded"
        elif track["download_status"] == "error":
            review_state = "download_failed"
        elif str(track["match_status"]).startswith("approved_"):
            review_state = "ready_to_download"
        elif track["match_status"] == "suggested":
            review_state = "review_match"
        elif track["match_status"] in {"match_error", "unmatched"}:
            review_state = "match_failed_or_unmatched"
        else:
            review_state = "needs_matching"
        source = (
            "liked_song + saved_album" if track["is_liked"] and track["is_saved_album"]
            else "liked_song" if track["is_liked"] else "saved_album"
        )
        rows.append({
            "spotify_id": track["spotify_id"],
            "added_at": track["added_at"] or "",
            "library_source": source,
            "title": track["title"],
            "artists": track["artists"],
            "album": track["album"] or "",
            "spotify_runtime_seconds": spotify_seconds,
            "match_status": track["match_status"],
            "youtube_score": track["youtube_score"] if track["youtube_score"] is not None else "",
            "youtube_title": track["youtube_title"] or "",
            "youtube_channel": track["youtube_channel"] or "",
            "youtube_runtime_seconds": youtube_seconds if youtube_seconds is not None else "",
            "runtime_difference_seconds": difference,
            "download_status": track["download_status"],
            "review_state": review_state,
            "spotify_url": track["spotify_url"] or "",
            "youtube_url": track["youtube_url"] or "",
        })
    return rows


def atomic_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    backup_existing_file(path)
    descriptor, name = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    os.close(descriptor)
    temporary = Path(name)
    try:
        with temporary.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.DictWriter(handle, fieldnames=REPORT_FIELDS)
            writer.writeheader()
            writer.writerows(rows)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def atomic_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise AppError(dependency_help("openpyxl")) from exc
    path.parent.mkdir(parents=True, exist_ok=True)
    backup_existing_file(path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Recent Additions"
    sheet.append(REPORT_FIELDS)
    for row in rows:
        sheet.append([row[field] for field in REPORT_FIELDS])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1DB954")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "A": 24, "B": 22, "C": 24, "D": 34, "E": 32, "F": 34,
        "H": 20, "I": 14, "J": 42, "K": 30, "N": 20, "O": 24,
        "P": 38, "Q": 38,
    }
    for column in range(1, len(REPORT_FIELDS) + 1):
        letter = get_column_letter(column)
        sheet.column_dimensions[letter].width = widths.get(letter, 18)
    for row_number in range(2, sheet.max_row + 1):
        for column in (16, 17):
            cell = sheet.cell(row_number, column)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
    descriptor, name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".xlsx", dir=path.parent
    )
    os.close(descriptor)
    temporary = Path(name)
    try:
        workbook.save(temporary)
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def run_stage(command: list[str], label: str) -> int:
    print(f"\n{label}")
    result = subprocess.run(command, check=False)
    if result.returncode:
        print(f"{label} completed with exit code {result.returncode}; continuing.")
    return result.returncode


def write_reports(args: argparse.Namespace, cutoff: str) -> list[dict[str, Any]]:
    with connect_db(args.db) as connection:
        rows = recent_rows(connection, cutoff)
    atomic_csv(args.report, rows)
    atomic_xlsx(args.xlsx, rows)
    print(f"Recent cutoff: {cutoff}")
    print(f"Recent tracks: {len(rows):,}")
    print(f"CSV: {args.report}")
    print(f"Excel: {args.xlsx}")
    states: dict[str, int] = {}
    for row in rows:
        states[row["review_state"]] = states.get(row["review_state"], 0) + 1
    for state in sorted(states):
        print(f"  {state}: {states[state]:,}")
    return rows


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        if not 0 <= args.auto_approve <= 100:
            raise AppError("--auto-approve must be between 0 and 100.")
        if not 0 <= args.min_score <= 100:
            raise AppError("--min-score must be between 0 and 100.")
        cutoff = cutoff_text(args)
        exit_code = 0
        if args.sync:
            exit_code |= run_stage(
                [
                    sys.executable,
                    str(PROJECT_DIR / "spotify_sync.py"),
                    "--db", str(args.db),
                    "--include-albums",
                ],
                "Refreshing Spotify liked songs and saved albums...",
            )
        rows = write_reports(args, cutoff)
        ids = [str(row["spotify_id"]) for row in rows]
        if args.match and ids:
            if args.refresh:
                match_ids = ids
            else:
                match_ids = [
                    str(row["spotify_id"]) for row in rows
                    if row["match_status"] in {
                        "not_searched", "match_error", "unmatched"
                    }
                ]
            if match_ids:
                command = [
                    sys.executable,
                    str(PROJECT_DIR / "youtube_match.py"),
                    "--db", str(args.db),
                    "--auto-approve", str(args.auto_approve),
                ]
                for spotify_id in match_ids:
                    command.extend(["--spotify-id", spotify_id])
                exit_code |= run_stage(command, "Scoring recent YouTube matches...")
            else:
                print("All recent tracks already have match results; no search needed.")
            rows = write_reports(args, cutoff)
        if args.download and ids:
            command = [
                sys.executable,
                str(PROJECT_DIR / "download_mp3.py"),
                "--db", str(args.db),
                "--min-score", str(args.min_score),
                "--all",
            ]
            if args.retry_errors:
                command.append("--retry-errors")
            if args.download_dry_run:
                command.append("--dry-run")
            if args.po_token_provider:
                command.append("--po-token-provider")
            for spotify_id in ids:
                command.extend(["--spotify-id", spotify_id])
            exit_code |= run_stage(command, "Downloading eligible recent tracks...")
            rows = write_reports(args, cutoff)
        if args.open_review:
            subprocess.run(["open", str(args.xlsx)], check=False)
        if not (args.sync or args.match or args.download):
            print("Report-only. Add --sync, --match, or --download for those stages.")
        return 1 if exit_code else 0
    except (AppError, OSError, subprocess.SubprocessError) as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
