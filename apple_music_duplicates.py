#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import unicodedata
import uuid
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from common import (
    AppError,
    DEFAULT_DB_PATH,
    PROJECT_DIR,
    backup_existing_file,
    connect_db,
    utc_now,
)
from activity_log import record_event


SOURCE_BRIDGE_PATH = PROJECT_DIR / "apple_music_bridge.applescript"
INSTALLED_BRIDGE_PATH = (
    Path(sys.prefix)
    / "share"
    / "spotify-youtube-library-mac"
    / "apple_music_bridge.applescript"
)
BRIDGE_PATH = (
    SOURCE_BRIDGE_PATH
    if SOURCE_BRIDGE_PATH.exists()
    else INSTALLED_BRIDGE_PATH
)
DEFAULT_REPORT = PROJECT_DIR / "data" / "apple_music_duplicate_report.csv"
DEFAULT_CLOSE_REPORT = PROJECT_DIR / "data" / "local_music_close_matches.xlsx"
DEFAULT_SWAP_REPORT = PROJECT_DIR / "data" / "apple_music_swaps.csv"
DEFAULT_PLAYLIST = "Spotify Archive Preferred"
DEFAULT_PLAN_DIR = PROJECT_DIR / "data" / "apple_music_restore_plans"
MARKER_PREFIX = "SPOTIFY_ARCHIVE_ID="
CLOSE_TITLE_SIMILARITY = 0.85
CLOSE_AUTO_PREFER_RUNTIME_SECONDS = 5.0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Compare downloaded MP3s with locally installed files in the Mac "
            "Music library and prefer exact title, album, artist, and runtime "
            "matches."
        )
    )
    parser.add_argument("--db", type=Path, default=DEFAULT_DB_PATH)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument(
        "--close-report",
        type=Path,
        default=DEFAULT_CLOSE_REPORT,
        help=(
            "Excel report for possible duplicates that are not changed. "
            f"Default: {DEFAULT_CLOSE_REPORT}"
        ),
    )
    parser.add_argument(
        "--swap-report",
        type=Path,
        default=DEFAULT_SWAP_REPORT,
        help=(
            "CSV list of exact Music entries replaced during --apply. "
            f"Default: {DEFAULT_SWAP_REPORT}"
        ),
    )
    parser.add_argument(
        "--duration-tolerance",
        type=float,
        default=5.0,
        help="Maximum runtime difference in seconds. Default: 5.",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help=(
            "Import preferred files and disable matching older Music entries. "
            "Without this flag, the command only creates a report."
        ),
    )
    parser.add_argument(
        "--import-new",
        action="store_true",
        help=(
            "With --apply, also import downloaded songs that have no Music "
            "duplicate."
        ),
    )
    parser.add_argument(
        "--spotify-id",
        action="append",
        default=[],
        metavar="ID",
        help=(
            "Compare or apply only this Spotify track ID. Repeat this option "
            "to target several downloaded tracks."
        ),
    )
    parser.add_argument("--limit", type=int)
    parser.add_argument("--playlist", default=DEFAULT_PLAYLIST)
    parser.add_argument(
        "--restore-plan",
        metavar="PLAN_ID",
        help=(
            "Restore enabled states and preferred-playlist membership recorded "
            "by a previous --apply plan. No library entry or file is deleted."
        ),
    )
    parser.add_argument(
        "--list-plans",
        action="store_true",
        help="List Apple Music restore plans without contacting Music.",
    )
    return parser.parse_args()


def normalize(value: str | None) -> str:
    text = unicodedata.normalize("NFKD", value or "")
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.casefold().replace("&", " and ")
    text = re.sub(r"[\W_]+", " ", text, flags=re.UNICODE)
    return " ".join(text.split())


def metadata_key(title: str, album: str, artist: str) -> tuple[str, str, str]:
    return normalize(title), normalize(album), normalize(artist)


def require_mac() -> None:
    if sys.platform != "darwin":
        raise AppError("This command must be run on a Mac.")
    if shutil.which("osascript") is None:
        raise AppError("macOS osascript was not found.")
    if not BRIDGE_PATH.exists():
        raise AppError(f"Apple Music helper is missing: {BRIDGE_PATH}")
    validate_bridge()


def validate_bridge() -> None:
    if shutil.which("osacompile") is None:
        raise AppError(
            "macOS osacompile was not found, so the Apple Music helper cannot "
            "be validated."
        )
    with tempfile.TemporaryDirectory(prefix="spotify-music-bridge-") as temp_dir:
        compiled = Path(temp_dir) / "bridge.scpt"
        result = subprocess.run(
            ["osacompile", "-o", str(compiled), str(BRIDGE_PATH)],
            text=True,
            capture_output=True,
            check=False,
            timeout=30,
        )
    if result.returncode != 0:
        detail = (result.stderr or result.stdout or "unknown compile error").strip()
        raise AppError(
            "The Apple Music AppleScript did not compile. No Music changes "
            f"were made: {detail[-1000:]}"
        )


def run_bridge(arguments: list[str], timeout: int = 1800) -> str:
    result = subprocess.run(
        ["osascript", str(BRIDGE_PATH), *arguments],
        text=True,
        capture_output=True,
        check=False,
        timeout=timeout,
    )
    if result.returncode != 0:
        message = (result.stderr or result.stdout or "Unknown Music error").strip()
        if "not authorized" in message.casefold() or "-1743" in message:
            raise AppError(
                "The app running this command does not have permission to "
                "control Music. Open "
                "System Settings > Privacy & Security > Automation, allow "
                "Terminal (or your current terminal app) to control Music, "
                "then run the command again. Report-only mode still needs "
                "read access through Automation."
            )
        raise AppError(f"Music automation failed: {message[-1200:]}")
    return result.stdout


def scan_music_library() -> list[dict[str, Any]]:
    print("Reading the Music library. A large library can take a few minutes...")
    output = run_bridge(["scan"])
    tracks = []
    for record in output.split("\x1e"):
        if not record:
            continue
        fields = record.split("\x1f")
        if len(fields) != 8:
            continue
        pid, title, artist, album, duration, enabled, location, comment = fields
        try:
            duration_value = float(duration)
        except (TypeError, ValueError):
            duration_value = None
        tracks.append(
            {
                "persistent_id": pid,
                "title": title,
                "artist": artist,
                "album": album,
                "duration": duration_value,
                "enabled": enabled.casefold() == "true",
                "location": location,
                "comment": comment,
            }
        )
    print(f"Read {len(tracks):,} Music library tracks.")
    return tracks


def mp3_duration(path: Path) -> float:
    try:
        from mutagen.mp3 import MP3
    except ImportError as exc:
        raise AppError(
            "mutagen is not installed. Activate .venv and run "
            "python -m pip install -r requirements.txt."
        ) from exc
    try:
        return float(MP3(path).info.length)
    except Exception as exc:
        raise AppError(f"Could not read MP3 runtime for {path}: {exc}") from exc


def marker_spotify_id(comment: str) -> str | None:
    match = re.search(r"\bSPOTIFY_ARCHIVE_ID=([A-Za-z0-9]+)", comment or "")
    return match.group(1) if match else None


def normalized_location(location: str | Path | None) -> str:
    """Return a stable key for comparing Music and download file paths."""
    if not location:
        return ""
    return os.path.normcase(os.path.realpath(os.fspath(location)))


def prefer_in_music(
    path: Path,
    spotify_id: str,
    playlist: str,
    title: str,
    artist: str,
    album: str,
    track_number: int,
    track_count: int,
    disc_number: int,
    compilation: bool,
    existing_path_id: str,
    old_ids: list[str],
) -> tuple[str, str, bool, bool, bool]:
    output = run_bridge(
        [
            "prefer",
            str(path.resolve()),
            spotify_id,
            playlist,
            title,
            artist,
            album,
            str(track_number),
            str(track_count),
            str(disc_number),
            "true" if compilation else "false",
            existing_path_id,
            *old_ids,
        ]
    ).strip()
    fields = output.split("\x1f")
    fields.extend(["", "false", "false", "false"])
    return (
        fields[0],
        fields[1],
        fields[2].casefold() == "true",
        fields[3].casefold() == "true",
        fields[4].casefold() == "true",
    )


def restore_in_music(plan: dict[str, Any]) -> None:
    old_tracks = json.loads(plan["old_tracks_json"])
    arguments = [
        "restore",
        plan["preferred_id"],
        "true" if plan["preferred_existed_before"] else "false",
        "true" if plan["preferred_enabled_before"] else "false",
        plan["playlist_name"],
        "true" if plan["playlist_had_track"] else "false",
    ]
    for item in old_tracks:
        arguments.extend(
            [
                item["persistent_id"],
                "true" if item["enabled"] else "false",
            ]
        )
    run_bridge(arguments)


def plan_path(plan_id: str, root: Path = DEFAULT_PLAN_DIR) -> Path:
    return root / f"{plan_id}.json"


def write_plan_file(plan: dict[str, Any], root: Path = DEFAULT_PLAN_DIR) -> Path:
    root.mkdir(parents=True, exist_ok=True)
    destination = plan_path(plan["plan_id"], root)
    temporary = destination.with_suffix(".json.tmp")
    temporary.write_text(
        json.dumps(plan, indent=2, ensure_ascii=False, sort_keys=True),
        encoding="utf-8",
    )
    os.replace(temporary, destination)
    return destination


def list_restore_plans(connection: Any) -> None:
    rows = connection.execute(
        """
        SELECT plan_id, spotify_id, action, status, created_at, applied_at,
               restored_at
        FROM apple_music_restore_plans
        ORDER BY created_at DESC
        """
    ).fetchall()
    if not rows:
        print("No Apple Music restore plans.")
        return
    for row in rows:
        print(
            f"{row['plan_id']}  {row['status']:<10}  "
            f"{row['spotify_id']}  {row['action']}  {row['created_at']}"
        )


def restore_saved_plan(connection: Any, plan_id: str, db_path: Path) -> None:
    row = connection.execute(
        "SELECT * FROM apple_music_restore_plans WHERE plan_id = ?",
        (plan_id,),
    ).fetchone()
    if not row:
        raise AppError(f"Unknown Apple Music restore plan: {plan_id}")
    plan = dict(row)
    if plan["status"] != "applied":
        raise AppError(
            f"Restore plan {plan_id} has status {plan['status']}; "
            "only applied plans can be restored."
        )
    restore_in_music(plan)
    restored_at = utc_now()
    connection.execute(
        """
        UPDATE apple_music_restore_plans
        SET status = 'restored', restored_at = ?, restore_error = NULL
        WHERE plan_id = ?
        """,
        (restored_at, plan_id),
    )
    connection.execute(
        """
        UPDATE tracks
        SET apple_music_status = 'restored',
            apple_music_updated_at = ?
        WHERE spotify_id = ?
        """,
        (restored_at, plan["spotify_id"]),
    )
    updated = dict(plan, status="restored", restored_at=restored_at)
    write_plan_file(updated, db_path.parent / "apple_music_restore_plans")
    record_event(
        connection,
        stage="apple_music",
        event="plan_restored",
        status="restored",
        spotify_id=plan["spotify_id"],
        details={"plan_id": plan_id},
        log_path=db_path.parent / "activity.jsonl",
    )
    connection.commit()
    print(f"Restored Apple Music plan: {plan_id}")


REPORT_COLUMNS = [
    "spotify_id",
    "title",
    "artist",
    "album",
    "download_path",
    "download_runtime",
    "metadata_matches",
    "eligible_runtime_matches",
    "protected_vinyl_matches",
    "eligible_non_mp3_matches",
    "eligible_close_runtime_matches",
    "closest_runtime_difference",
    "music_persistent_ids",
    "music_locations",
    "action",
]

CLOSE_REPORT_COLUMNS = [
    "spotify_id",
    "download_title",
    "download_artist",
    "download_album",
    "download_runtime_seconds",
    "download_path",
    "local_title",
    "local_artist",
    "local_album",
    "local_runtime_seconds",
    "runtime_difference_seconds",
    "title_exact",
    "artist_exact",
    "album_exact",
    "title_similarity",
    "reason",
    "local_enabled",
    "local_persistent_id",
    "local_file_location",
    "local_file_format",
    "action",
]

SWAP_REPORT_COLUMNS = [
    "spotify_id",
    "title",
    "artist",
    "album",
    "preferred_mp3",
    "old_music_persistent_id",
    "old_music_location",
    "old_music_format",
    "runtime_difference_seconds",
    "action",
    "restore_plan_id",
]


def write_csv_report(
    path: Path, columns: list[str], rows: list[dict[str, Any]]
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    backup_existing_file(path)
    temporary = path.with_suffix(f"{path.suffix}.{uuid.uuid4().hex}.tmp")
    try:
        with temporary.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.DictWriter(handle, fieldnames=columns)
            writer.writeheader()
            writer.writerows(rows)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def close_match_indexes(
    music_tracks: list[dict[str, Any]],
) -> dict[str, dict[Any, list[dict[str, Any]]]]:
    indexes: dict[str, dict[Any, list[dict[str, Any]]]] = {
        "title_artist": defaultdict(list),
        "title_album": defaultdict(list),
        "album_artist": defaultdict(list),
    }
    for item in music_tracks:
        title = normalize(item["title"])
        artist = normalize(item["artist"])
        album = normalize(item["album"])
        indexes["title_artist"][(title, artist)].append(item)
        indexes["title_album"][(title, album)].append(item)
        indexes["album_artist"][(album, artist)].append(item)
    return indexes


def vinyl_protected(album: str | None) -> bool:
    return bool(re.search(r"\(\s*vinyl\s*\)", album or "", re.IGNORECASE))


def local_file_format(location: str | None) -> str:
    suffix = Path(location or "").suffix.casefold().lstrip(".")
    return suffix or "unknown"


def close_match_rows(
    track: Any,
    download_path: Path,
    download_duration: float,
    indexes: dict[str, dict[Any, list[dict[str, Any]]]],
    duration_tolerance: float,
) -> list[dict[str, Any]]:
    title = normalize(track["title"])
    artist = normalize(track["artists"])
    album = normalize(track["album"])
    possible: dict[tuple[str, str], dict[str, Any]] = {}
    for index_name, key in (
        ("title_artist", (title, artist)),
        ("title_album", (title, album)),
        ("album_artist", (album, artist)),
    ):
        for item in indexes[index_name].get(key, []):
            identity = (
                str(item.get("persistent_id") or ""),
                str(item.get("location") or id(item)),
            )
            possible[identity] = item

    output: list[dict[str, Any]] = []
    for item in possible.values():
        local_title = normalize(item["title"])
        local_artist = normalize(item["artist"])
        local_album = normalize(item["album"])
        title_exact = local_title == title
        artist_exact = local_artist == artist
        album_exact = local_album == album
        title_similarity = SequenceMatcher(None, title, local_title).ratio()
        local_duration = item.get("duration")
        runtime_difference = (
            abs(float(local_duration) - download_duration)
            if local_duration is not None
            else None
        )

        if title_exact and artist_exact and album_exact:
            if (
                runtime_difference is not None
                and runtime_difference <= duration_tolerance
            ):
                # Exact eligible duplicates belong in the main report, not the
                # close/no-change workbook.
                continue
            reason = (
                "metadata_exact_runtime_unknown"
                if runtime_difference is None
                else "metadata_exact_runtime_outside_tolerance"
            )
        elif title_exact and artist_exact:
            reason = "title_artist_exact_album_differs"
        elif title_exact and album_exact:
            reason = "title_album_exact_artist_differs"
        elif album_exact and artist_exact and title_similarity >= CLOSE_TITLE_SIMILARITY:
            reason = "album_artist_exact_title_close"
        else:
            continue

        if (
            not vinyl_protected(item["album"])
            and runtime_difference is not None
            and runtime_difference < CLOSE_AUTO_PREFER_RUNTIME_SECONDS
        ):
            # This close candidate is promoted to the exact replacement flow
            # and will be recorded in the applied swap CSV instead.
            continue

        output.append(
            {
                "spotify_id": track["spotify_id"],
                "download_title": track["title"],
                "download_artist": track["artists"],
                "download_album": track["album"],
                "download_runtime_seconds": round(download_duration, 3),
                "download_path": str(download_path),
                "local_title": item["title"],
                "local_artist": item["artist"],
                "local_album": item["album"],
                "local_runtime_seconds": (
                    round(float(local_duration), 3)
                    if local_duration is not None
                    else None
                ),
                "runtime_difference_seconds": (
                    round(runtime_difference, 3)
                    if runtime_difference is not None
                    else None
                ),
                "title_exact": title_exact,
                "artist_exact": artist_exact,
                "album_exact": album_exact,
                "title_similarity": round(title_similarity, 3),
                "reason": reason,
                "local_enabled": bool(item["enabled"]),
                "local_persistent_id": item["persistent_id"],
                "local_file_location": item["location"],
                "local_file_format": local_file_format(item["location"]),
                "action": "review_only_no_change",
            }
        )
    return output


def close_runtime_replacement_candidates(
    track: Any,
    download_duration: float,
    indexes: dict[str, dict[Any, list[dict[str, Any]]]],
) -> list[dict[str, Any]]:
    title = normalize(track["title"])
    artist = normalize(track["artists"])
    album = normalize(track["album"])
    possible: dict[tuple[str, str], dict[str, Any]] = {}
    for index_name, key in (
        ("title_artist", (title, artist)),
        ("title_album", (title, album)),
        ("album_artist", (album, artist)),
    ):
        for item in indexes[index_name].get(key, []):
            identity = (
                str(item.get("persistent_id") or ""),
                str(item.get("location") or id(item)),
            )
            possible[identity] = item

    eligible: list[dict[str, Any]] = []
    for item in possible.values():
        if vinyl_protected(item["album"]) or item.get("duration") is None:
            continue
        local_title = normalize(item["title"])
        local_artist = normalize(item["artist"])
        local_album = normalize(item["album"])
        title_exact = local_title == title
        artist_exact = local_artist == artist
        album_exact = local_album == album
        if title_exact and artist_exact and album_exact:
            continue
        metadata_close = (
            (title_exact and artist_exact)
            or (title_exact and album_exact)
            or (
                album_exact
                and artist_exact
                and SequenceMatcher(None, title, local_title).ratio()
                >= CLOSE_TITLE_SIMILARITY
            )
        )
        difference = abs(float(item["duration"]) - download_duration)
        if metadata_close and difference < CLOSE_AUTO_PREFER_RUNTIME_SECONDS:
            eligible.append(item)
    return eligible


def write_close_report(
    path: Path,
    rows: list[dict[str, Any]],
    duration_tolerance: float,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise AppError(
            "openpyxl is not installed. Activate .venv and run: "
            "python -m pip install -r requirements.txt"
        ) from exc

    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instructions.append(["Local Music Library: Close Duplicate Review"])
    instructions.append(
        [
            "These rows are possible duplicates that were deliberately left "
            "unchanged. No local Music entry or audio file was deleted."
        ]
    )
    instructions.append(
        [
            "Automatic preference requires exact normalized title, album, and "
            f"artist plus a runtime difference of {duration_tolerance:g} "
            "seconds or less."
        ]
    )
    instructions.append(
        [
            "Close matches include exact metadata with a runtime mismatch, "
            "title+artist with a different album, title+album with a different "
            "artist, or a very similar title on the same artist and album."
        ]
    )
    instructions["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    instructions["A1"].fill = PatternFill("solid", fgColor="1DB954")
    instructions.column_dimensions["A"].width = 110
    for instruction_row in instructions.iter_rows():
        instruction_row[0].alignment = Alignment(
            wrap_text=True, vertical="top"
        )

    sheet = workbook.create_sheet("Close Matches")
    sheet.append(CLOSE_REPORT_COLUMNS)
    for item in rows:
        sheet.append([item.get(column) for column in CLOSE_REPORT_COLUMNS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 34
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="191414")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    widths = {
        "spotify_id": 24,
        "download_title": 34,
        "download_artist": 28,
        "download_album": 32,
        "download_runtime_seconds": 21,
        "download_path": 52,
        "local_title": 34,
        "local_artist": 28,
        "local_album": 32,
        "local_runtime_seconds": 20,
        "runtime_difference_seconds": 25,
        "title_similarity": 16,
        "reason": 42,
        "local_persistent_id": 22,
        "local_file_location": 52,
        "local_file_format": 18,
        "action": 24,
    }
    for index, column in enumerate(CLOSE_REPORT_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(
            column, 14
        )
    for column in (
        "download_runtime_seconds",
        "local_runtime_seconds",
        "runtime_difference_seconds",
    ):
        column_index = CLOSE_REPORT_COLUMNS.index(column) + 1
        for cells in sheet.iter_cols(
            min_col=column_index, max_col=column_index, min_row=2
        ):
            for cell in cells:
                cell.number_format = "0.000"
    similarity_index = CLOSE_REPORT_COLUMNS.index("title_similarity") + 1
    for cells in sheet.iter_cols(
        min_col=similarity_index, max_col=similarity_index, min_row=2
    ):
        for cell in cells:
            cell.number_format = "0.0%"

    path.parent.mkdir(parents=True, exist_ok=True)
    backup_existing_file(path)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".xlsx", dir=path.parent
    )
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, path)
    finally:
        temporary_path.unlink(missing_ok=True)


def main() -> int:
    args = parse_args()
    try:
        if args.list_plans:
            with connect_db(args.db) as connection:
                list_restore_plans(connection)
            return 0
        if args.restore_plan:
            require_mac()
            with connect_db(args.db) as connection:
                restore_saved_plan(connection, args.restore_plan, args.db)
            return 0
        require_mac()
        if args.duration_tolerance < 0:
            raise AppError("--duration-tolerance cannot be negative.")
        music_tracks = scan_music_library()
        by_metadata: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(
            list
        )
        by_marker: dict[str, list[dict[str, Any]]] = defaultdict(list)
        by_location: dict[str, dict[str, Any]] = {}
        for track in music_tracks:
            by_metadata[
                metadata_key(track["title"], track["album"], track["artist"])
            ].append(track)
            marked_id = marker_spotify_id(track["comment"])
            if marked_id:
                by_marker[marked_id].append(track)
            location_key = normalized_location(track.get("location"))
            if location_key:
                by_location[location_key] = track
        close_indexes = close_match_indexes(music_tracks)

        with connect_db(args.db) as connection:
            sql = """
                SELECT *
                FROM tracks
                WHERE (is_liked = 1 OR is_saved_album = 1)
                  AND user_deleted = 0
                  AND download_status = 'downloaded'
                  AND download_path IS NOT NULL
            """
            parameters: list[Any] = []
            if args.spotify_id:
                placeholders = ",".join("?" for _ in args.spotify_id)
                sql += f" AND spotify_id IN ({placeholders})"
                parameters.extend(args.spotify_id)
            sql += """
                ORDER BY primary_artist COLLATE NOCASE,
                         album COLLATE NOCASE,
                         disc_number,
                         track_number
            """
            if args.limit is not None:
                sql += " LIMIT ?"
                parameters.append(args.limit)
            downloaded = connection.execute(sql, parameters).fetchall()
            report_rows: list[dict[str, Any]] = []
            close_rows: list[dict[str, Any]] = []
            swap_rows: list[dict[str, Any]] = []
            progress_actions: dict[str, int] = defaultdict(int)

            for index, track in enumerate(downloaded, start=1):
                spotify_id = track["spotify_id"]
                path = Path(track["download_path"])
                print(
                    f"[{index:,}/{len(downloaded):,}] "
                    f"{track['primary_artist']} - {track['title']}"
                )
                row = {
                    "spotify_id": spotify_id,
                    "title": track["title"],
                    "artist": track["artists"],
                    "album": track["album"],
                    "download_path": str(path),
                    "download_runtime": "",
                    "metadata_matches": 0,
                    "eligible_runtime_matches": 0,
                    "protected_vinyl_matches": 0,
                    "eligible_non_mp3_matches": 0,
                    "eligible_close_runtime_matches": 0,
                    "closest_runtime_difference": "",
                    "music_persistent_ids": "",
                    "music_locations": "",
                    "action": "",
                }

                if not path.exists():
                    row["action"] = "download_missing"
                    report_rows.append(row)
                    continue

                preferred_existing = by_marker.get(spotify_id, [])
                existing_path_track = by_location.get(normalized_location(path))
                duration = mp3_duration(path)
                row["download_runtime"] = round(duration, 3)
                close_rows.extend(
                    close_match_rows(
                        track,
                        path,
                        duration,
                        close_indexes,
                        args.duration_tolerance,
                    )
                )
                close_eligible = close_runtime_replacement_candidates(
                    track, duration, close_indexes
                )
                row["eligible_close_runtime_matches"] = len(close_eligible)
                all_matches = by_metadata.get(
                    metadata_key(
                        track["title"], track["album"], track["artists"]
                    ),
                    [],
                )
                protected_matches = [
                    item for item in all_matches if vinyl_protected(item["album"])
                ]
                matches = [
                    item for item in all_matches if not vinyl_protected(item["album"])
                ]
                row["metadata_matches"] = len(matches)
                row["protected_vinyl_matches"] = len(protected_matches)
                differences = [
                    (
                        abs(float(item["duration"]) - duration),
                        item,
                    )
                    for item in matches
                    if item["duration"] is not None
                ]
                differences.sort(key=lambda pair: pair[0])
                eligible = [
                    item
                    for difference, item in differences
                    if difference <= args.duration_tolerance
                ]
                eligible_non_mp3 = [
                    item
                    for item in [*eligible, *close_eligible]
                    if local_file_format(item.get("location")) != "mp3"
                ]
                row["eligible_runtime_matches"] = len(eligible)
                row["eligible_non_mp3_matches"] = len(eligible_non_mp3)
                if differences:
                    row["closest_runtime_difference"] = round(
                        differences[0][0], 3
                    )
                row["music_persistent_ids"] = "; ".join(
                    item["persistent_id"] for item in matches
                )
                row["music_locations"] = "; ".join(
                    item["location"] for item in matches if item["location"]
                )

                if preferred_existing:
                    action = "already_preferred"
                elif close_eligible:
                    action = "would_prefer_close_match_under_5s"
                elif matches and not eligible:
                    action = "runtime_mismatch"
                elif eligible:
                    action = (
                        "would_prefer_mp3_over_other_format"
                        if eligible_non_mp3
                        else "would_prefer_download"
                    )
                else:
                    action = "would_import_new" if args.import_new else "no_match"

                should_apply = (
                    args.apply
                    and not preferred_existing
                    and (
                        bool(eligible)
                        or bool(close_eligible)
                        or (not matches and args.import_new)
                    )
                )
                if should_apply:
                    replacement_candidates = [*eligible, *close_eligible]
                    old_ids = [
                        item["persistent_id"] for item in replacement_candidates
                    ]
                    old_tracks = [
                        {
                            "persistent_id": item["persistent_id"],
                            "enabled": bool(item["enabled"]),
                            "location": item["location"],
                            "comment": item["comment"],
                        }
                        for item in replacement_candidates
                    ]
                    action = (
                        "preferred_close_match_under_5s"
                        if close_eligible
                        else (
                            "preferred_mp3_over_other_format"
                            if eligible_non_mp3
                            else (
                                "preferred_download"
                                if eligible or preferred_existing
                                else "imported_new"
                            )
                        )
                    )
                    plan_id = uuid.uuid4().hex
                    created_at = utc_now()
                    plan = {
                        "plan_id": plan_id,
                        "spotify_id": spotify_id,
                        "action": action,
                        "download_path": str(path.resolve()),
                        "playlist_name": args.playlist,
                        "old_tracks_json": json.dumps(old_tracks),
                        "old_tracks": old_tracks,
                        "preferred_id": None,
                        "preferred_location": None,
                        "preferred_existed_before": None,
                        "preferred_enabled_before": None,
                        "playlist_had_track": None,
                        "status": "planned",
                        "created_at": created_at,
                        "applied_at": None,
                        "restored_at": None,
                        "restore_error": None,
                    }
                    connection.execute(
                        """
                        INSERT INTO apple_music_restore_plans (
                            plan_id, spotify_id, action, download_path,
                            playlist_name, old_tracks_json, status, created_at
                        ) VALUES (?, ?, ?, ?, ?, ?, 'planned', ?)
                        """,
                        (
                            plan_id,
                            spotify_id,
                            action,
                            str(path.resolve()),
                            args.playlist,
                            json.dumps(old_tracks),
                            created_at,
                        ),
                    )
                    write_plan_file(
                        plan, args.db.parent / "apple_music_restore_plans"
                    )
                    connection.commit()
                    try:
                        (
                            preferred_id,
                            preferred_location,
                            preferred_existed_before,
                            preferred_enabled_before,
                            playlist_had_track,
                        ) = prefer_in_music(
                            path,
                            spotify_id,
                            args.playlist,
                            track["title"],
                            track["artists"],
                            track["album"],
                            int(track["track_number"] or 0),
                            int(track["total_tracks"] or 0),
                            int(track["disc_number"] or 0),
                            True,
                            (
                                str(existing_path_track["persistent_id"])
                                if existing_path_track
                                else ""
                            ),
                            old_ids,
                        )
                    except Exception as exc:
                        connection.execute(
                            """
                            UPDATE apple_music_restore_plans
                            SET status = 'apply_failed', restore_error = ?
                            WHERE plan_id = ?
                            """,
                            (str(exc)[:1000], plan_id),
                        )
                        write_plan_file(
                            dict(
                                plan,
                                status="apply_failed",
                                restore_error=str(exc)[:1000],
                            ),
                            args.db.parent / "apple_music_restore_plans",
                        )
                        connection.commit()
                        raise
                    applied_at = utc_now()
                    connection.execute(
                        """
                        UPDATE apple_music_restore_plans
                        SET preferred_id = ?,
                            preferred_location = ?,
                            preferred_existed_before = ?,
                            preferred_enabled_before = ?,
                            playlist_had_track = ?,
                            status = 'applied',
                            applied_at = ?
                        WHERE plan_id = ?
                        """,
                        (
                            preferred_id,
                            preferred_location,
                            1 if preferred_existed_before else 0,
                            1 if preferred_enabled_before else 0,
                            1 if playlist_had_track else 0,
                            applied_at,
                            plan_id,
                        ),
                    )
                    connection.execute(
                        """
                        UPDATE tracks
                        SET apple_music_status = ?,
                            apple_music_preferred_id = ?,
                            apple_music_replaced_ids = ?,
                            apple_music_updated_at = ?
                        WHERE spotify_id = ?
                        """,
                        (
                            action,
                            preferred_id,
                            json.dumps(old_ids),
                            applied_at,
                            spotify_id,
                        ),
                    )
                    write_plan_file(
                        dict(
                            plan,
                            preferred_id=preferred_id,
                            preferred_location=preferred_location,
                            preferred_existed_before=preferred_existed_before,
                            preferred_enabled_before=preferred_enabled_before,
                            playlist_had_track=playlist_had_track,
                            status="applied",
                            applied_at=applied_at,
                        ),
                        args.db.parent / "apple_music_restore_plans",
                    )
                    record_event(
                        connection,
                        stage="apple_music",
                        event="plan_applied",
                        status=action,
                        spotify_id=spotify_id,
                        details={"plan_id": plan_id, "old_ids": old_ids},
                        log_path=args.db.parent / "activity.jsonl",
                    )
                    connection.commit()
                    print(f"  Restore plan: {plan_id}")
                    for old_item in replacement_candidates:
                        old_duration = old_item.get("duration")
                        swap_rows.append(
                            {
                                "spotify_id": spotify_id,
                                "title": track["title"],
                                "artist": track["artists"],
                                "album": track["album"],
                                "preferred_mp3": str(path.resolve()),
                                "old_music_persistent_id": old_item[
                                    "persistent_id"
                                ],
                                "old_music_location": old_item["location"],
                                "old_music_format": local_file_format(
                                    old_item["location"]
                                ),
                                "runtime_difference_seconds": (
                                    round(
                                        abs(float(old_duration) - duration), 3
                                    )
                                    if old_duration is not None
                                    else None
                                ),
                                "action": action,
                                "restore_plan_id": plan_id,
                            }
                        )
                    if preferred_location:
                        row["music_locations"] = (
                            f"{row['music_locations']}; {preferred_location}"
                        ).strip("; ")
                else:
                    # Report-only mode must remain read-only so it can safely
                    # run while the downloader checkpoints another track.
                    if args.apply and action != "already_preferred":
                        connection.execute(
                            """
                            UPDATE tracks
                            SET apple_music_status = ?,
                                apple_music_updated_at = ?
                            WHERE spotify_id = ?
                            """,
                            (action, utc_now(), spotify_id),
                        )
                row["action"] = action
                report_rows.append(row)
                progress_actions[action] += 1
                if index % 25 == 0 or index == len(downloaded):
                    print(
                        "Comparison progress: "
                        f"{index:,}/{len(downloaded):,} "
                        f"({index / max(1, len(downloaded)):.1%}); "
                        f"swaps {len(swap_rows):,}; "
                        f"imports {progress_actions.get('imported_new', 0):,}; "
                        "protected VINYL "
                        f"{sum(int(item.get('protected_vinyl_matches') or 0) for item in report_rows):,}; "
                        f"review-only {len(close_rows):,}."
                    )

            connection.commit()

        write_csv_report(args.report, REPORT_COLUMNS, report_rows)
        write_close_report(
            args.close_report,
            close_rows,
            args.duration_tolerance,
        )

        counts: dict[str, int] = defaultdict(int)
        for row in report_rows:
            counts[row["action"]] += 1
        print(f"Report: {args.report}")
        print(
            f"Close-match Excel report: {args.close_report} "
            f"({len(close_rows):,} unchanged candidates)"
        )
        for action, count in sorted(counts.items()):
            print(f"  {action}: {count:,}")
        if args.apply:
            write_csv_report(args.swap_report, SWAP_REPORT_COLUMNS, swap_rows)
            print(
                f"Applied swap list: {args.swap_report} "
                f"({len(swap_rows):,} replaced Music entries)"
            )
        if not args.apply:
            print(
                "This was a report-only run. Review the CSV, then use --apply "
                "when you are ready."
            )
        return 0
    except (AppError, KeyboardInterrupt, subprocess.TimeoutExpired) as exc:
        print(f"\nError: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
