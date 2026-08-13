from __future__ import annotations

import csv
import io
import sys
import tempfile
import unittest
from contextlib import redirect_stdout
from pathlib import Path
from unittest.mock import patch

import apple_music_duplicates
from common import connect_db, utc_now


class AppleMusicWorkflowTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.db = self.root / "music.sqlite"
        self.mp3 = self.root / "song.mp3"
        self.mp3.write_bytes(b"test")
        with connect_db(self.db) as connection:
            connection.execute(
                """
                INSERT INTO tracks (
                    spotify_id, title, artists, primary_artist, album,
                    duration_ms, explicit, is_liked, match_status,
                    download_status, download_path, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "spotify123",
                    "Test Song",
                    "Test Artist",
                    "Test Artist",
                    "Test Album",
                    180000,
                    0,
                    1,
                    "approved_manual",
                    "downloaded",
                    str(self.mp3),
                    utc_now(),
                ),
            )

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def music_tracks(self) -> list[dict[str, object]]:
        return [
            {
                "persistent_id": "OLD123",
                "title": "Test Song",
                "artist": "Test Artist",
                "album": "Test Album",
                "duration": 182.0,
                "enabled": True,
                "location": "/Music/old.mp3",
                "comment": "",
            }
        ]

    def test_report_finds_runtime_safe_duplicate(self) -> None:
        report = self.root / "report.csv"
        close_report = self.root / "close.xlsx"
        argv = [
            "apple_music_duplicates.py",
            "--db",
            str(self.db),
            "--report",
            str(report),
            "--close-report",
            str(close_report),
        ]
        with (
            patch.object(apple_music_duplicates, "require_mac"),
            patch.object(
                apple_music_duplicates,
                "scan_music_library",
                return_value=self.music_tracks(),
            ),
            patch.object(
                apple_music_duplicates, "mp3_duration", return_value=180.0
            ),
            patch.object(sys, "argv", argv),
            redirect_stdout(io.StringIO()),
        ):
            self.assertEqual(apple_music_duplicates.main(), 0)
        with report.open(encoding="utf-8-sig") as handle:
            row = next(csv.DictReader(handle))
        self.assertEqual(row["action"], "would_prefer_download")
        self.assertEqual(row["eligible_runtime_matches"], "1")
        self.assertEqual(row["closest_runtime_difference"], "2.0")
        with connect_db(self.db) as connection:
            status = connection.execute(
                "SELECT apple_music_status FROM tracks "
                "WHERE spotify_id = 'spotify123'"
            ).fetchone()[0]
        self.assertEqual(status, "not_checked")
        from openpyxl import load_workbook

        workbook = load_workbook(close_report, read_only=True)
        self.assertEqual(workbook["Close Matches"].max_row, 1)

    def test_close_runtime_mismatch_is_written_to_excel_and_not_applied(
        self,
    ) -> None:
        report = self.root / "report.csv"
        close_report = self.root / "close.xlsx"
        local_track = dict(self.music_tracks()[0], duration=190.0)
        argv = [
            "apple_music_duplicates.py",
            "--db",
            str(self.db),
            "--report",
            str(report),
            "--close-report",
            str(close_report),
        ]
        with (
            patch.object(apple_music_duplicates, "require_mac"),
            patch.object(
                apple_music_duplicates,
                "scan_music_library",
                return_value=[local_track],
            ),
            patch.object(
                apple_music_duplicates, "mp3_duration", return_value=180.0
            ),
            patch.object(sys, "argv", argv),
            redirect_stdout(io.StringIO()),
        ):
            self.assertEqual(apple_music_duplicates.main(), 0)

        from openpyxl import load_workbook

        workbook = load_workbook(close_report, read_only=True, data_only=True)
        sheet = workbook["Close Matches"]
        headings = [cell.value for cell in sheet[1]]
        values = [cell.value for cell in sheet[2]]
        row = dict(zip(headings, values, strict=True))
        self.assertEqual(
            row["reason"], "metadata_exact_runtime_outside_tolerance"
        )
        self.assertEqual(row["runtime_difference_seconds"], 10.0)
        self.assertEqual(row["action"], "review_only_no_change")

    def test_report_explicitly_prefers_downloaded_mp3_over_m4a(self) -> None:
        report = self.root / "report.csv"
        close_report = self.root / "close.xlsx"
        local_track = dict(self.music_tracks()[0], location="/Music/old.m4a")
        argv = [
            "apple_music_duplicates.py", "--db", str(self.db),
            "--report", str(report), "--close-report", str(close_report),
        ]
        with (
            patch.object(apple_music_duplicates, "require_mac"),
            patch.object(
                apple_music_duplicates,
                "scan_music_library",
                return_value=[local_track],
            ),
            patch.object(
                apple_music_duplicates, "mp3_duration", return_value=180.0
            ),
            patch.object(sys, "argv", argv),
            redirect_stdout(io.StringIO()),
        ):
            self.assertEqual(apple_music_duplicates.main(), 0)
        with report.open(encoding="utf-8-sig") as handle:
            row = next(csv.DictReader(handle))
        self.assertEqual(
            row["action"], "would_prefer_mp3_over_other_format"
        )
        self.assertEqual(row["eligible_non_mp3_matches"], "1")

    def test_close_metadata_match_under_five_seconds_is_preferred(self) -> None:
        report = self.root / "report.csv"
        swaps = self.root / "swaps.csv"
        local_track = dict(
            self.music_tracks()[0],
            album="Different Album",
            duration=184.999,
            location="/Music/old.m4a",
        )
        argv = [
            "apple_music_duplicates.py", "--db", str(self.db),
            "--report", str(report),
            "--close-report", str(self.root / "close.xlsx"),
            "--swap-report", str(swaps), "--apply",
        ]
        with (
            patch.object(apple_music_duplicates, "require_mac"),
            patch.object(
                apple_music_duplicates,
                "scan_music_library",
                return_value=[local_track],
            ),
            patch.object(
                apple_music_duplicates, "mp3_duration", return_value=180.0
            ),
            patch.object(
                apple_music_duplicates,
                "prefer_in_music",
                return_value=("NEW456", "/Music/new.mp3", False, False, False),
            ) as prefer,
            patch.object(sys, "argv", argv),
            redirect_stdout(io.StringIO()),
        ):
            self.assertEqual(apple_music_duplicates.main(), 0)
        self.assertEqual(prefer.call_args.args[-1], ["OLD123"])
        self.assertEqual(prefer.call_args.args[-2], "")
        with report.open(encoding="utf-8-sig") as handle:
            row = next(csv.DictReader(handle))
        self.assertEqual(row["action"], "preferred_close_match_under_5s")
        self.assertEqual(row["eligible_close_runtime_matches"], "1")
        with swaps.open(encoding="utf-8-sig") as handle:
            swap = next(csv.DictReader(handle))
        self.assertEqual(swap["old_music_format"], "m4a")

    def test_close_match_at_exactly_five_seconds_is_not_preferred(self) -> None:
        local_track = dict(
            self.music_tracks()[0],
            album="Different Album",
            duration=185.0,
        )
        indexes = apple_music_duplicates.close_match_indexes([local_track])
        track = {
            "title": "Test Song",
            "artists": "Test Artist",
            "album": "Test Album",
        }
        candidates = (
            apple_music_duplicates.close_runtime_replacement_candidates(
                track, 180.0, indexes
            )
        )
        self.assertEqual(candidates, [])

    def test_apply_records_preferred_music_id(self) -> None:
        with connect_db(self.db) as connection:
            connection.execute(
                """
                UPDATE tracks SET track_number = 3, total_tracks = 12,
                                  disc_number = 2
                WHERE spotify_id = 'spotify123'
                """
            )
        report = self.root / "report.csv"
        close_report = self.root / "close.xlsx"
        argv = [
            "apple_music_duplicates.py",
            "--db",
            str(self.db),
            "--report",
            str(report),
            "--close-report",
            str(close_report),
            "--apply",
            "--swap-report",
            str(self.root / "swaps.csv"),
        ]
        with (
            patch.object(apple_music_duplicates, "require_mac"),
            patch.object(
                apple_music_duplicates,
                "scan_music_library",
                return_value=self.music_tracks(),
            ),
            patch.object(
                apple_music_duplicates, "mp3_duration", return_value=180.0
            ),
            patch.object(
                apple_music_duplicates,
                "prefer_in_music",
                return_value=(
                    "NEW456",
                    "/Music/new.mp3",
                    False,
                    False,
                    False,
                ),
            ) as prefer,
            patch.object(sys, "argv", argv),
            redirect_stdout(io.StringIO()),
        ):
            self.assertEqual(apple_music_duplicates.main(), 0)
        prefer.assert_called_once()
        self.assertEqual(prefer.call_args.args[-6:-2], (3, 12, 2, True))
        with connect_db(self.db) as connection:
            track = connection.execute(
                """
                SELECT apple_music_status, apple_music_preferred_id
                FROM tracks
                WHERE spotify_id = 'spotify123'
                """
            ).fetchone()
        self.assertEqual(track["apple_music_status"], "preferred_download")
        self.assertEqual(track["apple_music_preferred_id"], "NEW456")
        with connect_db(self.db) as connection:
            plan = connection.execute(
                "SELECT * FROM apple_music_restore_plans"
            ).fetchone()
        self.assertEqual(plan["status"], "applied")
        self.assertEqual(plan["preferred_id"], "NEW456")
        self.assertTrue(
            (
                self.root
                / "apple_music_restore_plans"
                / f"{plan['plan_id']}.json"
            ).exists()
        )

    def test_vinyl_album_entry_is_never_disabled_or_replaced(self) -> None:
        with connect_db(self.db) as connection:
            connection.execute(
                "UPDATE tracks SET album = 'Test Album (VINYL)' "
                "WHERE spotify_id = 'spotify123'"
            )
        local_track = dict(self.music_tracks()[0], album="Test Album (VINYL)")
        report = self.root / "report.csv"
        close_report = self.root / "close.xlsx"
        argv = [
            "apple_music_duplicates.py",
            "--db", str(self.db),
            "--report", str(report),
            "--close-report", str(close_report),
            "--apply", "--import-new",
            "--swap-report", str(self.root / "swaps.csv"),
        ]
        with (
            patch.object(apple_music_duplicates, "require_mac"),
            patch.object(
                apple_music_duplicates,
                "scan_music_library",
                return_value=[local_track],
            ),
            patch.object(
                apple_music_duplicates, "mp3_duration", return_value=180.0
            ),
            patch.object(
                apple_music_duplicates,
                "prefer_in_music",
                return_value=("NEW456", "/Music/new.mp3", False, False, False),
            ) as prefer,
            patch.object(sys, "argv", argv),
            redirect_stdout(io.StringIO()),
        ):
            self.assertEqual(apple_music_duplicates.main(), 0)
        self.assertEqual(prefer.call_args.args[-1], [])
        with report.open(encoding="utf-8-sig") as handle:
            row = next(csv.DictReader(handle))
        self.assertEqual(row["protected_vinyl_matches"], "1")
        self.assertEqual(row["action"], "imported_new")

    def test_apply_reuses_download_already_imported_by_path(self) -> None:
        local_track = dict(
            self.music_tracks()[0],
            persistent_id="IMPORTED123",
            location=str(self.mp3.resolve()),
            comment="",
        )
        argv = [
            "apple_music_duplicates.py", "--db", str(self.db),
            "--report", str(self.root / "report.csv"),
            "--close-report", str(self.root / "close.xlsx"),
            "--swap-report", str(self.root / "swaps.csv"),
            "--apply",
        ]
        with (
            patch.object(apple_music_duplicates, "require_mac"),
            patch.object(
                apple_music_duplicates,
                "scan_music_library",
                return_value=[local_track],
            ),
            patch.object(
                apple_music_duplicates, "mp3_duration", return_value=180.0
            ),
            patch.object(
                apple_music_duplicates,
                "prefer_in_music",
                return_value=(
                    "IMPORTED123", str(self.mp3.resolve()), True, True, False
                ),
            ) as prefer,
            patch.object(sys, "argv", argv),
            redirect_stdout(io.StringIO()),
        ):
            self.assertEqual(apple_music_duplicates.main(), 0)
        self.assertEqual(prefer.call_args.args[-2], "IMPORTED123")

    def test_apply_skips_track_already_marked_as_preferred(self) -> None:
        marked = dict(
            self.music_tracks()[0],
            persistent_id="NEW456",
            location=str(self.mp3),
            comment="SPOTIFY_ARCHIVE_ID=spotify123",
        )
        report = self.root / "report.csv"
        argv = [
            "apple_music_duplicates.py", "--db", str(self.db),
            "--report", str(report),
            "--close-report", str(self.root / "close.xlsx"),
            "--swap-report", str(self.root / "swaps.csv"),
            "--apply", "--import-new",
        ]
        with (
            patch.object(apple_music_duplicates, "require_mac"),
            patch.object(
                apple_music_duplicates,
                "scan_music_library",
                return_value=[marked],
            ),
            patch.object(
                apple_music_duplicates, "mp3_duration", return_value=180.0
            ),
            patch.object(apple_music_duplicates, "prefer_in_music") as prefer,
            patch.object(sys, "argv", argv),
            redirect_stdout(io.StringIO()),
        ):
            self.assertEqual(apple_music_duplicates.main(), 0)
        prefer.assert_not_called()
        with report.open(encoding="utf-8-sig") as handle:
            row = next(csv.DictReader(handle))
        self.assertEqual(row["action"], "already_preferred")

    def test_restore_plan_records_reversal_without_deleting_library_entry(
        self,
    ) -> None:
        with connect_db(self.db) as connection:
            connection.execute(
                """
                INSERT INTO apple_music_restore_plans (
                    plan_id, spotify_id, action, download_path, playlist_name,
                    old_tracks_json, preferred_id, preferred_existed_before,
                    preferred_enabled_before, playlist_had_track, status,
                    created_at, applied_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "plan123",
                    "spotify123",
                    "preferred_download",
                    str(self.mp3),
                    "Spotify Archive Preferred",
                    '[{"persistent_id":"OLD123","enabled":true}]',
                    "NEW456",
                    0,
                    0,
                    0,
                    "applied",
                    utc_now(),
                    utc_now(),
                ),
            )
        argv = [
            "apple_music_duplicates.py",
            "--db",
            str(self.db),
            "--restore-plan",
            "plan123",
        ]
        with (
            patch.object(apple_music_duplicates, "require_mac"),
            patch.object(
                apple_music_duplicates, "restore_in_music"
            ) as restore,
            patch.object(sys, "argv", argv),
            redirect_stdout(io.StringIO()),
        ):
            self.assertEqual(apple_music_duplicates.main(), 0)
        restore.assert_called_once()
        with connect_db(self.db) as connection:
            plan = connection.execute(
                """
                SELECT status, restored_at
                FROM apple_music_restore_plans
                WHERE plan_id = 'plan123'
                """
            ).fetchone()
        self.assertEqual(plan["status"], "restored")
        self.assertIsNotNone(plan["restored_at"])


if __name__ == "__main__":
    unittest.main()
