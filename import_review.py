#!/usr/bin/env python3
from __future__ import annotations

import argparse
import uuid
from pathlib import Path
from typing import Any

from common import (
    AppError,
    DEFAULT_CSV_PATH,
    DEFAULT_DB_PATH,
    DEFAULT_XLSX_PATH,
    connect_db,
    dependency_help,
    export_catalog,
    utc_now,
)
from activity_log import record_event
from matching_rules import SCORING_VERSION


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import approvals and manual URLs from the Excel review file."
    )
    parser.add_argument("--db", type=Path, default=DEFAULT_DB_PATH)
    parser.add_argument("--csv", type=Path, default=DEFAULT_CSV_PATH)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX_PATH)
    return parser.parse_args()


def youtube_info(url: str) -> dict[str, Any]:
    try:
        import yt_dlp
    except ImportError as exc:
        raise AppError(dependency_help("yt-dlp")) from exc
    options = {
        "quiet": True,
        "no_warnings": True,
        "skip_download": True,
        "noplaylist": True,
        "socket_timeout": 20,
    }
    try:
        with yt_dlp.YoutubeDL(options) as ydl:
            info = ydl.extract_info(url, download=False)
    except Exception as exc:
        raise AppError(f"Could not validate manual YouTube URL: {exc}") from exc
    return {
        "youtube_url": info.get("webpage_url") or url,
        "youtube_video_id": info.get("id"),
        "youtube_title": info.get("title") or "",
        "youtube_channel": info.get("channel") or info.get("uploader") or "",
        "youtube_channel_verified": 1
        if info.get("channel_is_verified") or info.get("uploader_is_verified")
        else 0,
        "youtube_duration_seconds": info.get("duration"),
    }


def record_review_decision(
    connection: Any,
    spotify_id: str,
    decision: str,
    log_path: Path,
) -> None:
    track = connection.execute(
        "SELECT * FROM tracks WHERE spotify_id = ?", (spotify_id,)
    ).fetchone()
    if not track:
        return
    timestamp = utc_now()
    connection.execute(
        """
        INSERT INTO match_assessments (
            run_id, spotify_id, youtube_video_id, youtube_url,
            youtube_title, youtube_channel, youtube_duration_seconds,
            score, hard_reject, automatic_approval_eligible,
            approval_threshold, decision, reason, score_notes,
            scoring_version, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, 0, NULL, ?, ?, ?, ?, ?)
        """,
        (
            uuid.uuid4().hex,
            spotify_id,
            track["youtube_video_id"],
            track["youtube_url"],
            track["youtube_title"],
            track["youtube_channel"],
            track["youtube_duration_seconds"],
            track["youtube_score"],
            decision,
            "Decision imported from the review workbook",
            track["youtube_score_notes"],
            int(track["youtube_score_version"] or SCORING_VERSION),
            timestamp,
        ),
    )
    record_event(
        connection,
        stage="approval",
        event="review_imported",
        status=decision,
        spotify_id=spotify_id,
        details={"youtube_url": track["youtube_url"]},
        log_path=log_path,
    )


def main() -> int:
    args = parse_args()
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(
            "Error: openpyxl is not installed. Run: "
            "python -m pip install -r requirements.txt"
        )
        return 1
    if not args.xlsx.exists():
        print(f"Error: review workbook not found: {args.xlsx}")
        return 1

    workbook = load_workbook(args.xlsx, data_only=True, read_only=True)
    if "Tracks" not in workbook.sheetnames:
        print("Error: the workbook does not contain a Tracks sheet.")
        return 1
    sheet = workbook["Tracks"]
    headers = {
        str(cell.value).strip(): index
        for index, cell in enumerate(next(sheet.iter_rows(values_only=False)))
        if cell.value
    }
    required = {"decision", "manual_youtube_url", "spotify_id"}
    if not required.issubset(headers):
        print("Error: the Tracks sheet is missing required review columns.")
        return 1

    decisions = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        decision = str(values[headers["decision"]] or "").strip().casefold()
        if not decision:
            continue
        decisions.append(
            {
                "decision": decision,
                "manual_url": str(
                    values[headers["manual_youtube_url"]] or ""
                ).strip(),
                "spotify_id": str(values[headers["spotify_id"]] or "").strip(),
            }
        )
    workbook.close()

    try:
        with connect_db(args.db) as connection:
            changed = 0
            for item in decisions:
                spotify_id = item["spotify_id"]
                decision = item["decision"]
                existing = connection.execute(
                    "SELECT * FROM tracks WHERE spotify_id = ?", (spotify_id,)
                ).fetchone()
                if not existing:
                    print(f"Skipping unknown Spotify ID: {spotify_id}")
                    continue
                if decision == "approve":
                    if item["manual_url"]:
                        info = youtube_info(item["manual_url"])
                        source_changed = (
                            existing["youtube_url"] != info["youtube_url"]
                        )
                        connection.execute(
                            """
                            UPDATE tracks
                            SET youtube_url = :youtube_url,
                                youtube_video_id = :youtube_video_id,
                                youtube_title = :youtube_title,
                                youtube_channel = :youtube_channel,
                                youtube_channel_verified =
                                    :youtube_channel_verified,
                                youtube_duration_seconds =
                                    :youtube_duration_seconds,
                                youtube_score = NULL,
                                youtube_score_notes = 'Manually selected URL',
                                youtube_score_version = :score_version,
                                match_status = 'approved_manual',
                                reviewed_at = :reviewed_at,
                                download_status = CASE
                                    WHEN :source_changed
                                    THEN 'not_downloaded'
                                    ELSE download_status
                                END,
                                download_path = CASE
                                    WHEN :source_changed THEN NULL
                                    ELSE download_path
                                END,
                                download_error = CASE
                                    WHEN :source_changed THEN NULL
                                    ELSE download_error
                                END,
                                download_error_code = CASE
                                    WHEN :source_changed THEN NULL
                                    ELSE download_error_code
                                END,
                                download_next_retry_at = CASE
                                    WHEN :source_changed THEN NULL
                                    ELSE download_next_retry_at
                                END,
                                downloaded_at = CASE
                                    WHEN :source_changed THEN NULL
                                    ELSE downloaded_at
                                END,
                                downloaded_duration_seconds = CASE
                                    WHEN :source_changed THEN NULL
                                    ELSE downloaded_duration_seconds
                                END,
                                downloaded_duration_difference_seconds = CASE
                                    WHEN :source_changed THEN NULL
                                    ELSE downloaded_duration_difference_seconds
                                END,
                                apple_music_status = CASE
                                    WHEN :source_changed THEN 'not_checked'
                                    ELSE apple_music_status
                                END,
                                apple_music_preferred_id = CASE
                                    WHEN :source_changed THEN NULL
                                    ELSE apple_music_preferred_id
                                END,
                                apple_music_replaced_ids = CASE
                                    WHEN :source_changed THEN NULL
                                    ELSE apple_music_replaced_ids
                                END,
                                apple_music_updated_at = CASE
                                    WHEN :source_changed THEN NULL
                                    ELSE apple_music_updated_at
                                END
                            WHERE spotify_id = :spotify_id
                            """,
                            dict(
                                info,
                                spotify_id=spotify_id,
                                reviewed_at=utc_now(),
                                source_changed=source_changed,
                                score_version=SCORING_VERSION,
                            ),
                        )
                    else:
                        selected = connection.execute(
                            "SELECT youtube_url FROM tracks WHERE spotify_id = ?",
                            (spotify_id,),
                        ).fetchone()
                        if not selected or not selected["youtube_url"]:
                            print(
                                f"Cannot approve {spotify_id}: no suggested or "
                                "manual YouTube URL."
                            )
                            continue
                        connection.execute(
                            """
                            UPDATE tracks
                            SET match_status = 'approved_manual',
                                reviewed_at = ?
                            WHERE spotify_id = ?
                            """,
                            (utc_now(), spotify_id),
                        )
                    changed += 1
                    record_review_decision(
                        connection,
                        spotify_id,
                        "approved_manual",
                        args.db.parent / "activity.jsonl",
                    )
                elif decision == "skip":
                    connection.execute(
                        """
                        UPDATE tracks
                        SET match_status = 'skipped', reviewed_at = ?
                        WHERE spotify_id = ?
                        """,
                        (utc_now(), spotify_id),
                    )
                    changed += 1
                    record_review_decision(
                        connection,
                        spotify_id,
                        "skipped",
                        args.db.parent / "activity.jsonl",
                    )
                elif decision == "reset":
                    connection.execute(
                        """
                        UPDATE tracks
                        SET match_status = CASE
                                WHEN youtube_url IS NULL THEN 'not_searched'
                                ELSE 'suggested'
                            END,
                            reviewed_at = NULL
                        WHERE spotify_id = ?
                        """,
                        (spotify_id,),
                    )
                    changed += 1
                    record_review_decision(
                        connection,
                        spotify_id,
                        "reset",
                        args.db.parent / "activity.jsonl",
                    )
                else:
                    print(
                        f"Ignoring decision '{decision}' for {spotify_id}. "
                        "Use approve, skip, or reset."
                    )
            connection.commit()
            export_catalog(connection, args.csv, args.xlsx)
        print(f"Imported {changed:,} review decision(s).")
        print(f"Updated workbook: {args.xlsx}")
        return 0
    except (AppError, KeyboardInterrupt) as exc:
        print(f"\nError: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
